import streamlit as st
import plotly.graph_objects as go
import numpy as np
import io
import re
import requests
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from datetime import datetime

# ─────────────────────────────────────────────
# AIRTABLE CONFIG
# ─────────────────────────────────────────────

AIRTABLE_TOKEN   = st.secrets.get("AIRTABLE_TOKEN", "")
AIRTABLE_BASE_ID = st.secrets.get("AIRTABLE_BASE_ID", "")
AIRTABLE_URL     = f"https://api.airtable.com/v0/{AIRTABLE_BASE_ID}"

def save_lead(table, fields):
    try:
        requests.post(
            f"{AIRTABLE_URL}/{table}",
            headers={
                "Authorization": f"Bearer {AIRTABLE_TOKEN}",
                "Content-Type": "application/json"
            },
            json={"fields": fields},
            timeout=5
        )
    except Exception:
        pass

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, PageBreak, Image as RLImage
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.lib.enums import TA_JUSTIFY, TA_CENTER

# ─────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────

st.set_page_config(
    page_title="Freedom Simulator | The Freedom Project",
    page_icon="freedom_symbol_purple_white.svg",
    layout="wide"
)

# ─────────────────────────────────────────────
# DESIGN SYSTEM
# ─────────────────────────────────────────────

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@400;600;700&family=Inter:wght@400;500&display=swap');

html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
.stApp { background-color: #0B0B0F; color: #F2F2F2; }
h1, h2, h3 { font-family: 'Space Grotesk', sans-serif; }

div[data-testid="stMetric"] {
    background: #0f0f16;
    border-radius: 12px;
    padding: 20px;
    border: 0.5px solid rgba(255,255,255,0.08);
}
div[data-testid="stMetric"] label { color: #71717A !important; font-size: 12px !important; }
div[data-testid="stMetric"] [data-testid="stMetricValue"] {
    color: #F2F2F2 !important;
    font-family: 'Space Grotesk', sans-serif !important;
}

.stButton > button {
    background: #5A54C4; color: white; border-radius: 9px; border: none;
    padding: 12px 28px; font-family: 'Inter', sans-serif; font-weight: 500;
    font-size: 14px; width: 100%;
}
.stButton > button:hover { background: #4840a8; border: none; }

.stDownloadButton > button {
    background: #5A54C4 !important;
    color: white !important;
    border-radius: 9px !important;
    border: none !important;
    padding: 12px 28px !important;
    font-family: 'Inter', sans-serif !important;
    font-weight: 500 !important;
    font-size: 14px !important;
    width: 100% !important;
}

.stDownloadButton > button:hover {
    background: #4840a8 !important;
    border: none !important;
}
.stSelectbox > div > div { background: #0f0f16; border: 0.5px solid rgba(255,255,255,0.1); border-radius: 8px; }
.stNumberInput > div > div > input { background: #0f0f16; border: 0.5px solid rgba(255,255,255,0.1); border-radius: 8px; color: #F2F2F2; }
.stTextInput > div > div > input { background: #0f0f16; border: 0.5px solid rgba(255,255,255,0.1); border-radius: 8px; color: #F2F2F2; }
.stSlider > div > div > div > div { background: #5A54C4; }

/* Hide streamlit branding */
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding-top: 2rem; padding-bottom: 2rem; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def card(content, border_color="rgba(255,255,255,0.08)", bg="#0f0f16", padding="28px 24px"):
    return f"""<div style="background:{bg};border:0.5px solid {border_color};border-radius:14px;padding:{padding};margin-bottom:16px">{content}</div>"""

def label(text):
    return f'<div style="font-size:11px;letter-spacing:1.4px;text-transform:uppercase;color:#3f3f46;font-weight:500;margin-bottom:8px">{text}</div>'

def divider(margin="24px 0"):
    return f'<div style="border-top:0.5px solid rgba(255,255,255,0.07);margin:{margin}"></div>'

# ─────────────────────────────────────────────
# PDF GENERATOR
# ─────────────────────────────────────────────

def generate_pdf(inp, score, inv_needed, max_cap, target_cap, fn, fi_age, survives, gap,
                 mc_prob=None, coast_age=None, contrib_pct=None, returns_pct=None,
                 total_contrib=None, total_returns=None):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches

    buffer = io.BytesIO()

    # ── COLORS ──
    C_BG       = colors.HexColor("#0B0B0F")
    C_SURFACE  = colors.HexColor("#0f0f16")
    C_SURFACE2 = colors.HexColor("#141420")
    C_PURPLE   = colors.HexColor("#5A54C4")
    C_PURPLE_L = colors.HexColor("#7F77DD")
    C_PURPLE_D = colors.HexColor("#26215C")
    C_YELLOW   = colors.HexColor("#f9f09d")
    C_WHITE    = colors.HexColor("#F2F2F2")
    C_MUTED    = colors.HexColor("#71717A")
    C_SUBTLE   = colors.HexColor("#3f3f46")
    C_LINE     = colors.HexColor("#1f1f2e")
    C_SUCCESS  = colors.HexColor("#4ade80")
    C_WARN     = colors.HexColor("#EF9F27")
    C_ERROR    = colors.HexColor("#E24B4A")

    PW, PH = A4
    ML = MR = 1.8*cm
    MT = MB = 1.8*cm
    W  = PW - ML - MR

    # ── STYLES ──
    def ps(name, **kw):
        return ParagraphStyle(name, **kw)

    sty_brand    = ps("brand",   fontName="Helvetica-Bold", fontSize=9,  textColor=C_PURPLE,   spaceAfter=0,  leading=12, letterSpacing=1.5)
    sty_cover_h  = ps("coverh",  fontName="Helvetica-Bold", fontSize=38, textColor=C_WHITE,    spaceAfter=8,  leading=44, letterSpacing=-0.5)
    sty_cover_s  = ps("covers",  fontName="Helvetica",      fontSize=11, textColor=C_MUTED,    spaceAfter=0,  leading=16)
    sty_h1       = ps("h1",      fontName="Helvetica-Bold", fontSize=16, textColor=C_WHITE,    spaceAfter=6,  leading=20, spaceBefore=20)
    sty_h2       = ps("h2",      fontName="Helvetica-Bold", fontSize=11, textColor=C_WHITE,    spaceAfter=4,  leading=14, spaceBefore=14)
    sty_body     = ps("body",    fontName="Helvetica",      fontSize=9,  textColor=C_MUTED,    spaceAfter=6,  leading=14, alignment=TA_JUSTIFY)
    sty_body_w   = ps("bodyw",   fontName="Helvetica",      fontSize=9,  textColor=C_WHITE,    spaceAfter=6,  leading=14)
    sty_label    = ps("label",   fontName="Helvetica-Bold", fontSize=7,  textColor=C_SUBTLE,   spaceAfter=2,  leading=10, letterSpacing=1.2)
    sty_value    = ps("value",   fontName="Helvetica-Bold", fontSize=18, textColor=C_WHITE,    spaceAfter=0,  leading=22)
    sty_caption  = ps("cap",     fontName="Helvetica",      fontSize=7,  textColor=C_SUBTLE,   spaceAfter=0,  leading=10)
    sty_insight  = ps("ins",     fontName="Helvetica",      fontSize=9,  textColor=C_MUTED,    spaceAfter=8,  leading=14, leftIndent=12)

    score_hex = "#7F77DD" if score >= 70 else "#EF9F27" if score >= 40 else "#E24B4A"
    C_SCORE   = colors.HexColor(score_hex)
    score_txt = "Strong" if score >= 80 else "On track" if score >= 60 else "Needs attention" if score >= 40 else "At risk"

    r_name    = inp.get("name", "")
    r_email   = inp.get("email", "")
    r_country = inp.get("country", "")
    mi        = inp.get("monthly_investment", 0)
    mc_inc    = inp.get("monthly_income", 0)
    ya        = inp.get("years_accumulation", 0)
    yr        = inp.get("years_retirement", 0)
    ar        = inp.get("ann_return", 0.06)
    profile   = inp.get("profile", "Balanced")
    age       = inp.get("current_age", 30)

    # ── CHART ──
    def make_chart():
        def sim(inv):
            cap, hist, mr = 0, [], ar/12
            for _ in range(ya*12):
                cap = cap*(1+mr)+inv
                hist.append(cap)
            for _ in range(yr*12):
                cap = cap*(1+mr)-mc_inc
                hist.append(max(cap,0))
                if cap <= 0: break
            return hist

        h_cur = sim(mi)
        h_rec = sim(inv_needed)
        x_cur = [i/12 for i in range(len(h_cur))]
        x_rec = [i/12 for i in range(len(h_rec))]

        fig, ax = plt.subplots(figsize=(7.2, 3.2))
        fig.patch.set_facecolor("#0f0f16")
        ax.set_facecolor("#0f0f16")

        ax.fill_between(x_cur, h_cur, alpha=0.15, color="#7F77DD")
        ax.plot(x_cur, h_cur, color="#7F77DD", linewidth=2.0, label="Your plan")
        ax.plot(x_rec, h_rec, color="#f9f09d", linewidth=1.4, linestyle="--", label="Recommended plan")
        ax.axvline(x=ya, color="#26215C", linewidth=1.2, linestyle=":")
        ax.text(ya+0.3, max(h_cur)*0.96, "Retirement", color="#3f3f46", fontsize=7)

        ax.set_xlabel("Years", color="#71717A", fontsize=8)
        ax.set_ylabel("Portfolio value (€)", color="#71717A", fontsize=8)
        ax.tick_params(colors="#3f3f46", labelsize=7)
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"€{int(x/1000)}k" if x >= 1000 else f"€{int(x)}"))
        for spine in ax.spines.values():
            spine.set_edgecolor("#1f1f2e")
        ax.grid(axis="y", color="#1f1f2e", linewidth=0.5, alpha=0.8)
        ax.grid(axis="x", visible=False)

        legend = ax.legend(fontsize=7, facecolor="#141420", edgecolor="#1f1f2e", labelcolor="#A1A1AA", loc="upper left")

        plt.tight_layout(pad=0.4)
        buf = io.BytesIO()
        plt.savefig(buf, format="png", dpi=180, facecolor="#0f0f16")
        plt.close()
        buf.seek(0)
        return buf

    chart_buf = make_chart()

    # ── CANVAS CALLBACKS ──
    def cover_page(canvas, doc):
        canvas.saveState()
        canvas.setFillColor(C_BG)
        canvas.rect(0, 0, PW, PH, fill=1, stroke=0)
        canvas.setFillColor(C_PURPLE)
        canvas.rect(0, 0, 0.35*cm, PH, fill=1, stroke=0)
        canvas.setFillColor(C_SURFACE)
        canvas.rect(0, PH-1.8*cm, PW, 1.8*cm, fill=1, stroke=0)
        # logo in header
        try:
            import os
            logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logo.png")
            if os.path.exists(logo_path):
                canvas.drawImage(logo_path, ML+0.5*cm, PH-1.55*cm, width=0.9*cm, height=0.9*cm, preserveAspectRatio=True, mask='auto')
            canvas.setFillColor(C_WHITE)
            canvas.setFont("Helvetica-Bold", 8)
            canvas.drawString(ML+1.6*cm, PH-1.1*cm, "THE FREEDOM PROJECT  (fi)")
        except:
            canvas.setFillColor(C_WHITE)
            canvas.setFont("Helvetica-Bold", 8)
            canvas.drawString(ML+0.5*cm, PH-1.1*cm, "THE FREEDOM PROJECT  (fi)")
        canvas.setFillColor(C_SUBTLE)
        canvas.setFont("Helvetica", 8)
        canvas.drawRightString(PW-MR, PH-1.1*cm, datetime.now().strftime("%B %Y"))
        cx, cy, r = PW-MR-2.5*cm, PH*0.58, 2.0*cm
        canvas.setFillColor(C_SURFACE)
        canvas.circle(cx, cy, r, fill=1, stroke=0)
        canvas.setStrokeColor(C_SCORE)
        canvas.setLineWidth(4)
        canvas.circle(cx, cy, r, fill=0, stroke=1)
        canvas.setFillColor(C_SCORE)
        canvas.setFont("Helvetica-Bold", 28)
        canvas.drawCentredString(cx, cy-0.3*cm, str(int(score)))
        canvas.setFillColor(C_SUBTLE)
        canvas.setFont("Helvetica", 7)
        canvas.drawCentredString(cx, cy-0.9*cm, "OUT OF 100")
        canvas.setFillColor(C_SCORE)
        canvas.setFont("Helvetica-Bold", 9)
        canvas.drawCentredString(cx, cy+0.6*cm, score_txt.upper())
        canvas.setStrokeColor(C_PURPLE_D)
        canvas.setLineWidth(0.5)
        canvas.line(ML+0.5*cm, PH*0.35, PW-MR, PH*0.35)
        canvas.setFillColor(C_SUBTLE)
        canvas.setFont("Helvetica", 7)
        canvas.drawString(ML+0.5*cm, MB*0.6, f"Prepared for {r_name}  ·  {r_email}  ·  {r_country}")
        canvas.drawRightString(PW-MR, MB*0.6, "For planning purposes only")
        canvas.restoreState()

    def inner_page(canvas, doc):
        canvas.saveState()
        canvas.setFillColor(C_BG)
        canvas.rect(0, 0, PW, PH, fill=1, stroke=0)
        canvas.setFillColor(C_PURPLE)
        canvas.rect(0, 0, 0.25*cm, PH, fill=1, stroke=0)
        # header
        canvas.setFillColor(C_SURFACE)
        canvas.rect(0, PH-1.2*cm, PW, 1.2*cm, fill=1, stroke=0)
        canvas.setFillColor(C_WHITE)
        canvas.setFont("Helvetica-Bold", 7)
        canvas.drawString(ML+0.3*cm, PH-0.75*cm, "THE FREEDOM PROJECT  (fi)")
        canvas.setFillColor(C_SUBTLE)
        canvas.setFont("Helvetica", 7)
        canvas.drawRightString(PW-MR, PH-0.75*cm, "Financial Independence Report")
        # footer
        canvas.setStrokeColor(C_LINE)
        canvas.setLineWidth(0.5)
        canvas.line(ML+0.3*cm, MB+0.3*cm, PW-MR, MB+0.3*cm)
        canvas.setFillColor(C_SUBTLE)
        canvas.setFont("Helvetica", 6.5)
        canvas.drawString(ML+0.3*cm, MB*0.5, f"{r_name}  ·  {r_country}  ·  {datetime.now().strftime('%B %d, %Y')}")
        canvas.drawRightString(PW-MR, MB*0.5, f"Page {doc.page}")
        canvas.restoreState()

    # ── DOCUMENT ──
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=ML+0.5*cm, rightMargin=MR,
        topMargin=2.2*cm, bottomMargin=1.8*cm
    )

    story = []

    # ════════════════════════════════
    # PAGE 1 — COVER
    # ════════════════════════════════
    story.append(Spacer(1, 3.5*cm))
    story.append(Paragraph("FINANCIAL INDEPENDENCE", ps("tag", fontName="Helvetica-Bold", fontSize=8, textColor=C_PURPLE, spaceAfter=6, letterSpacing=2)))
    story.append(Paragraph("Your Personal<br/>Report", sty_cover_h))
    story.append(Spacer(1, 0.4*cm))
    story.append(Paragraph(f"Based on a <b>{profile}</b> investment profile over <b>{ya} years</b> of accumulation.", sty_cover_s))
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph(f"Target retirement income: <b>€{mc_inc:,}/month</b>  ·  Freedom Number: <b>€{int(fn):,}</b>", sty_cover_s))
    story.append(PageBreak())

    # ════════════════════════════════
    # PAGE 2 — EXECUTIVE SUMMARY
    # ════════════════════════════════
    story.append(Paragraph("01 — Executive Summary", sty_h1))
    story.append(HRFlowable(width=W, color=C_PURPLE_D, thickness=0.5, spaceAfter=14))

    # Summary text
    gap_text = f"Your plan is on track. Your current monthly investment of €{mi:,} exceeds the minimum required." if gap == 0 else \
               f"Your current monthly investment of €{mi:,} falls €{gap:,} short of the recommended €{inv_needed:,}/month."
    fi_text  = f"At your current savings rate, you could reach financial independence at approximately age {fi_age}." if fi_age else \
               "Increasing your monthly investment would allow you to reach financial independence within your planning horizon."
    sust_text = f"Your portfolio sustains your desired income of €{mc_inc:,}/month for the full {yr}-year retirement period." if survives else \
                f"Your portfolio may be depleted before the end of your {yr}-year retirement. Consider increasing contributions."

    story.append(Paragraph(gap_text, sty_body))
    story.append(Paragraph(fi_text,  sty_body))
    story.append(Paragraph(sust_text, sty_body))
    story.append(Spacer(1, 0.5*cm))

    # 4-metric row — simple single-table approach
    metric_items = [
        ("READINESS SCORE",  f"{int(score)} / 100",                    score_hex),
        ("FREEDOM NUMBER",   f"€{int(fn):,}",                          "#7F77DD"),
        ("MONTHLY GAP",      "On track ✓" if gap==0 else f"+€{gap:,}", "#4ade80" if gap==0 else "#EF9F27"),
        ("FI AGE ESTIMATE",  f"Age {fi_age}" if fi_age else "—",       "#F2F2F2"),
    ]
    m_row = []
    for lbl, val, col in metric_items:
        m_row.append(Paragraph(
            f'<font size="6" color="#3f3f46">{lbl}</font><br/><br/>'
            f'<font size="15" color="{col}"><b>{val}</b></font>',
            ps(f"mc_{lbl}", fontName="Helvetica", fontSize=15, leading=20, spaceAfter=0)
        ))

    t_m = Table([m_row], colWidths=[W/4]*4)
    t_m.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,-1), C_SURFACE),
        ('LEFTPADDING',   (0,0), (-1,-1), 14),
        ('RIGHTPADDING',  (0,0), (-1,-1), 8),
        ('TOPPADDING',    (0,0), (-1,-1), 16),
        ('BOTTOMPADDING', (0,0), (-1,-1), 16),
        ('LINEAFTER',     (0,0), (2,0),   0.5, C_LINE),
        ('VALIGN',        (0,0), (-1,-1), 'TOP'),
    ]))
    story.append(t_m)
    story.append(Spacer(1, 0.6*cm))

    # Full metrics table
    story.append(Paragraph("02 — Key Metrics", sty_h1))
    story.append(HRFlowable(width=W, color=C_PURPLE_D, thickness=0.5, spaceAfter=14))

    all_metrics = [
        ("Monthly investment",       f"€{mi:,}"),
        ("Recommended investment",   f"€{inv_needed:,}"),
        ("Target retirement capital",f"€{int(target_cap):,}"),
        ("Maximum capital reached",  f"€{int(max_cap):,}"),
        ("Freedom Number",           f"€{int(fn):,}"),
        ("Estimated FI age",         f"Age {fi_age}" if fi_age else "—"),
        ("Investment profile",       profile),
        ("Annual return assumed",    f"{int(ar*100)}% (nominal)"),
        ("Accumulation period",      f"{ya} years"),
        ("Retirement period",        f"{yr} years"),
        ("Retirement sustained",     "Yes" if survives else "No"),
        ("Projection type",          "Nominal (inflation not adjusted)"),
    ]

    rows = [all_metrics[i:i+2] for i in range(0, len(all_metrics), 2)]
    t_rows = []
    for i, row in enumerate(rows):
        while len(row) < 2: row.append(("",""))
        bg = C_SURFACE if i % 2 == 0 else C_SURFACE2
        t_rows.append([
            Paragraph(f'<font color="#3f3f46">{row[0][0]}</font>', ps("tr", fontName="Helvetica", fontSize=8.5, textColor=C_MUTED, leading=12)),
            Paragraph(f'<b>{row[0][1]}</b>', ps("tv", fontName="Helvetica-Bold", fontSize=8.5, textColor=C_WHITE, leading=12)),
            Paragraph(f'<font color="#3f3f46">{row[1][0]}</font>', ps("tr2", fontName="Helvetica", fontSize=8.5, textColor=C_MUTED, leading=12)),
            Paragraph(f'<b>{row[1][1]}</b>', ps("tv2", fontName="Helvetica-Bold", fontSize=8.5, textColor=C_WHITE, leading=12)),
        ])

    t_full = Table(t_rows, colWidths=[W*0.28, W*0.22, W*0.28, W*0.22])
    style_rows = [
        ('LEFTPADDING',  (0,0), (-1,-1), 12),
        ('RIGHTPADDING', (0,0), (-1,-1), 12),
        ('TOPPADDING',   (0,0), (-1,-1), 9),
        ('BOTTOMPADDING',(0,0), (-1,-1), 9),
        ('LINEAFTER',    (1,0), (1,-1),  0.5, C_LINE),
        ('VALIGN',       (0,0), (-1,-1), 'MIDDLE'),
    ]
    for i in range(len(t_rows)):
        bg = C_SURFACE if i % 2 == 0 else C_SURFACE2
        style_rows.append(('BACKGROUND', (0,i), (-1,i), bg))
    t_full.setStyle(TableStyle(style_rows))
    story.append(t_full)
    story.append(PageBreak())

    # ════════════════════════════════
    # PAGE 3 — PROJECTION CHART
    # ════════════════════════════════
    story.append(Paragraph("03 — Portfolio Projection", sty_h1))
    story.append(HRFlowable(width=W, color=C_PURPLE_D, thickness=0.5, spaceAfter=10))
    story.append(Paragraph(
        f"The chart below shows your portfolio trajectory over {ya+yr} years under your current plan (purple) "
        f"versus the recommended investment of €{inv_needed:,}/month (yellow dashed). "
        f"The dotted vertical line marks the start of retirement at year {ya}.",
        sty_body
    ))
    story.append(Spacer(1, 0.3*cm))
    story.append(RLImage(chart_buf, width=W, height=W*0.44))
    story.append(Spacer(1, 0.2*cm))
    story.append(Paragraph(f"Assumed annual return: {int(ar*100)}%  ·  Profile: {profile}  ·  Nominal terms only", sty_caption))
    story.append(Spacer(1, 0.6*cm))

    # ════════════════════════════════
    # PAGE 3 continued — INSIGHTS
    # ════════════════════════════════
    story.append(Paragraph("04 — Personalised Insights", sty_h1))
    story.append(HRFlowable(width=W, color=C_PURPLE_D, thickness=0.5, spaceAfter=10))

    insights = []
    if gap > 0:
        insights.append(("Investment gap", f"Your current monthly investment of €{mi:,} is €{gap:,} below the recommended €{inv_needed:,}. "
            "Closing this gap is the single highest-impact action available to you. Even a partial increase improves long-term outcomes significantly."))
    else:
        insights.append(("On track", f"Your current monthly investment of €{mi:,} exceeds the minimum required to meet your retirement goal. "
            "You have a positive savings buffer that adds resilience to market downturns."))

    if fi_age:
        insights.append(("Financial independence", f"Based on your {profile.lower()} profile ({int(ar*100)}% annual return), "
            f"you could reach your Freedom Number of €{int(fn):,} at approximately age {fi_age}. "
            "This is the point at which your investments generate enough income to cover your expenses indefinitely."))

    insights.append(("Retirement sustainability", "Your plan sustains your desired retirement income for the full period." if survives else
        f"Your portfolio may be depleted before the end of your {yr}-year retirement. "
        "Consider increasing monthly contributions, reducing desired retirement income, or extending your accumulation period."))

    insights.append(("Inflation note", f"Projections are shown in nominal terms. A 2.5% annual inflation rate would reduce "
        f"the real purchasing power of your €{mc_inc:,}/month retirement income over time. "
        "Factor this into your planning by targeting a higher nominal income or investing in inflation-hedged assets."))

    insights.append(("Compounding effect", f"Of your projected maximum portfolio of €{int(max_cap):,}, "
        f"approximately {max(0, round((1 - (mi*12*ya)/max(max_cap,1))*100))}% is generated by investment returns rather than contributions. "
        "This illustrates the power of long-term compound growth."))

    for title_ins, text_ins in insights:
        t_ins = Table([[
            Paragraph(
                f'<font size="7" color="#7F77DD"><b>{title_ins.upper()}</b></font><br/><br/>'
                f'<font size="9" color="#71717A">{text_ins}</font>',
                ps(f"ins_{title_ins}", fontName="Helvetica", fontSize=9, leading=14, spaceAfter=0)
            )
        ]], colWidths=[W])
        t_ins.setStyle(TableStyle([
            ('BACKGROUND',    (0,0), (-1,-1), C_SURFACE),
            ('LEFTPADDING',   (0,0), (-1,-1), 16),
            ('RIGHTPADDING',  (0,0), (-1,-1), 16),
            ('TOPPADDING',    (0,0), (-1,-1), 12),
            ('BOTTOMPADDING', (0,0), (-1,-1), 12),
            ('LINEBEFORE',    (0,0), (0,-1),  2, colors.HexColor("#5A54C4")),
        ]))
        story.append(t_ins)
        story.append(Spacer(1, 5))

    # ── PRO SECTIONS ──
    if mc_prob is not None:
        story.append(PageBreak())
        story.append(Paragraph("05 — Pro Analysis", sty_h1))
        story.append(HRFlowable(width=W, color=C_PURPLE_D, thickness=0.5, spaceAfter=14))

        story.append(Paragraph("Monte Carlo Simulation", sty_h2))
        mc_color_pdf = colors.HexColor("#7F77DD") if mc_prob >= 70 else colors.HexColor("#EF9F27") if mc_prob >= 50 else colors.HexColor("#E24B4A")
        mc_row = [[
            Paragraph(f'<font size="40"><b>{mc_prob}%</b></font><br/><font size="9" color="#71717A">success rate</font>',
                      ps("mc_n", fontName="Helvetica-Bold", fontSize=40, textColor=mc_color_pdf, leading=46, alignment=TA_CENTER)),
            Paragraph(
                f"Based on 1,000 simulations with randomised annual returns, your plan has a <b>{mc_prob}% probability</b> of "
                f"sustaining €{inp.get('monthly_income',0):,}/month for {inp.get('years_retirement',25)} years. "
                f"{'Strong result.' if mc_prob >= 70 else 'Consider increasing contributions.' if mc_prob >= 50 else 'Significant adjustments recommended.'}",
                body
            )
        ]]
        t_mc = Table(mc_row, colWidths=[W*0.2, W*0.8])
        t_mc.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BACKGROUND',(0,0),(-1,-1),dark),
                                   ('LEFTPADDING',(0,0),(-1,-1),12),('RIGHTPADDING',(0,0),(-1,-1),12),
                                   ('TOPPADDING',(0,0),(-1,-1),12),('BOTTOMPADDING',(0,0),(-1,-1),12)]))
        story.append(t_mc)
        story.append(Spacer(1, 14))

        if contrib_pct is not None:
            story.append(Paragraph("Portfolio Breakdown — Contributions vs Returns", sty_h2))
            bd_row = [[
                Paragraph(f'YOUR CONTRIBUTIONS<br/><font size="18"><b>€{total_contrib:,}</b></font><br/><font size="9" color="#71717A">{contrib_pct}% of final portfolio</font>',
                          ps("b1", fontName="Helvetica", fontSize=9, textColor=colors.HexColor("#AFA9EC"), leading=20)),
                Paragraph(f'MARKET RETURNS<br/><font size="18"><b>€{total_returns:,}</b></font><br/><font size="9" color="#71717A">{returns_pct}% of final portfolio</font>',
                          ps("b2", fontName="Helvetica", fontSize=9, textColor=colors.HexColor("#7F77DD"), leading=20))
            ]]
            t_bd = Table(bd_row, colWidths=[W/2, W/2])
            t_bd.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),dark),
                                       ('LEFTPADDING',(0,0),(-1,-1),14),('RIGHTPADDING',(0,0),(-1,-1),14),
                                       ('TOPPADDING',(0,0),(-1,-1),12),('BOTTOMPADDING',(0,0),(-1,-1),12),
                                       ('LINEAFTER',(0,0),(0,-1),0.5,C_LINE)]))
            story.append(t_bd)
            story.append(Spacer(1, 14))

        if coast_age:
            story.append(Paragraph("Coast FI", sty_h2))
            story.append(Paragraph(
                f"You reach Coast FI at approximately age <b>{coast_age}</b>. At that point you could stop all contributions "
                f"and compounding alone would grow your portfolio to €{int(fn):,} by retirement, "
                f"without investing another euro.",
                body
            ))
            story.append(Spacer(1, 10))

    story.append(PageBreak())

    # ════════════════════════════════
    # PAGE 4 — DISCLAIMER
    # ════════════════════════════════
    story.append(Spacer(1, 1*cm))
    story.append(Paragraph("A Note on Methodology", sty_h1))
    story.append(HRFlowable(width=W, color=C_PURPLE_D, thickness=0.5, spaceAfter=14))
    story.append(Paragraph(
        "This report was built with one goal in mind: to give you a clear, honest picture of where you stand financially. "
        "The projections are based on mathematical models that assume consistent returns and contributions over time. "
        "Markets do not always behave consistently — but the underlying logic of compound growth and long-term investing does not change.",
        sty_body
    ))
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph(
        f"All projections use nominal returns of {int(ar*100)}% annually ({profile} profile) and do not account for inflation, "
        "taxes or platform fees. Real outcomes will vary. This report is a planning tool, not a prediction. "
        "Use it as a starting point for your financial thinking — and revisit it as your situation evolves.",
        sty_body
    ))
    story.append(Spacer(1, 0.8*cm))
    story.append(HRFlowable(width=W, color=C_LINE, thickness=0.5, spaceAfter=14))
    story.append(Paragraph(
        "THE FREEDOM PROJECT  (fi)  ·  Captain Compound  ·  financial independence",
        ps("sig", fontName="Helvetica-Bold", fontSize=8, textColor=C_PURPLE, leading=12, alignment=TA_CENTER)
    ))
    story.append(Spacer(1, 0.2*cm))
    story.append(Paragraph(
        f"Report generated {datetime.now().strftime('%B %d, %Y')}  ·  {r_name}  ·  {r_country}",
        ps("sig2", fontName="Helvetica", fontSize=7, textColor=C_SUBTLE, leading=10, alignment=TA_CENTER)
    ))

    # ── BUILD ──
    def get_template(is_cover):
        return cover_page if is_cover else inner_page

    doc.build(
        story,
        onFirstPage=cover_page,
        onLaterPages=inner_page
    )
    buffer.seek(0)
    return buffer


# ─────────────────────────────────────────────
# HEADER
# ─────────────────────────────────────────────

col_logo, col_text, col_right = st.columns([1, 8, 3])
with col_logo:
    try:
        st.image("freedom_symbol_purple_white.svg", width=44)
    except:
        st.markdown("<div style='width:44px'></div>", unsafe_allow_html=True)
with col_text:
    st.markdown("""
    <div style="padding-top:4px">
      <span style="font-family:'Space Grotesk',sans-serif;font-weight:700;font-size:17px;letter-spacing:1px;color:#F2F2F2">THE FREEDOM PROJECT</span>
      <span style="color:#9490e8;font-family:'Space Grotesk',sans-serif;font-weight:700;font-size:17px"> (fi)</span>
      <div style="font-size:11px;color:#3f3f46;letter-spacing:.5px;margin-top:2px">Freedom Simulator</div>
    </div>
    """, unsafe_allow_html=True)
with col_right:
    st.markdown("<div style='font-size:12px;color:#3f3f46;text-align:right;padding-top:10px'>financial independence</div>", unsafe_allow_html=True)

st.markdown("<div style='border-bottom:0.5px solid rgba(255,255,255,0.07);margin-bottom:32px;margin-top:12px'></div>", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────

for key, default in [
    ("step", 0), ("inputs", {}), ("results", None),
    ("show_pdf_gate", False), ("pdf_unlocked", False),
    ("pro_unlocked", False), ("pro_email", ""), ("pro_gate_open", False)
]:
    if key not in st.session_state:
        st.session_state[key] = default

# ─────────────────────────────────────────────
# STEP INDICATOR
# ─────────────────────────────────────────────

def step_indicator(current, total=3):
    steps = ["Your situation", "Your goal", "Investment profile"]
    cols = st.columns(total)
    for i, (col, s) in enumerate(zip(cols, steps)):
        n = i + 1
        if n < current:
            dot_bg, dot_color, text_color, dot_text = "rgba(127,119,221,0.2)", "#7F77DD", "#7F77DD", "✓"
        elif n == current:
            dot_bg, dot_color, text_color, dot_text = "rgba(127,119,221,0.15)", "#F2F2F2", "#F2F2F2", str(n)
        else:
            dot_bg, dot_color, text_color, dot_text = "transparent", "#3f3f46", "#3f3f46", str(n)
        with col:
            st.markdown(f"""
            <div style="text-align:center">
              <div style="width:30px;height:30px;border-radius:50%;background:{dot_bg};border:1px solid {dot_color};
                display:inline-flex;align-items:center;justify-content:center;font-size:12px;
                font-weight:600;color:{dot_color};margin-bottom:6px">{dot_text}</div>
              <div style="font-size:11px;color:{text_color};letter-spacing:.3px">{s}</div>
            </div>""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# SIMULATION FUNCTIONS
# ─────────────────────────────────────────────

def get_return(profile):
    return {"Conservative": 0.04, "Balanced": 0.06, "Aggressive": 0.08}[profile]

def simulate(monthly_inv, monthly_inc, years_acc, years_ret, ann_return):
    capital, history, mr = 0, [], ann_return / 12
    for _ in range(years_acc * 12):
        capital = capital * (1 + mr) + monthly_inv
        history.append(capital)
    survives = True
    for _ in range(years_ret * 12):
        capital = capital * (1 + mr) - monthly_inc
        history.append(capital)
        if capital <= 0:
            survives = False
            break
    return history, survives

def required_investment(monthly_inc, years_acc, years_ret, ann_return):
    invest = 50
    while invest < 20000:
        _, ok = simulate(invest, monthly_inc, years_acc, years_ret, ann_return)
        if ok: return invest
        invest += 25
    return invest

def freedom_number(monthly_exp, swr=0.04):
    return (monthly_exp * 12) / swr

def years_to_fi(monthly_inv, fn, ann_return):
    capital, mr = 0, ann_return / 12
    for m in range(1, 600):
        capital = capital * (1 + mr) + monthly_inv
        if capital >= fn: return m / 12
    return None

def monte_carlo(mi, mc, ya, yr, ar, n=1000):
    ok = sum(
        simulate(mi, mc, ya, yr, ar + np.random.normal(0, 0.015))[1]
        for _ in range(n)
    )
    return round(ok / n * 100, 1)

# ─────────────────────────────────────────────
# STEPS 1–3
# ─────────────────────────────────────────────

# ── STEP 0 — INTRO ──
if st.session_state.step == 0:
    st.markdown("""
    <div style="max-width:580px;margin:40px auto 0;text-align:center">
      <div style="display:inline-block;background:rgba(90,84,196,0.12);border:0.5px solid rgba(90,84,196,0.3);
           border-radius:20px;color:#9490e8;font-size:10px;letter-spacing:1px;text-transform:uppercase;
           padding:4px 14px;margin-bottom:24px">Freedom Simulator</div>
      <div style="font-family:'Space Grotesk',sans-serif;font-size:36px;font-weight:700;
           letter-spacing:-.5px;line-height:1.15;margin-bottom:16px">
        Find the number that<br>sets you free.
      </div>
      <div style="font-size:15px;color:#71717A;line-height:1.75;margin-bottom:40px;max-width:460px;margin-left:auto;margin-right:auto">
        This simulator calculates your <strong style="color:#A1A1AA">Freedom Number</strong> —
        the exact portfolio size at which your investments generate enough passive income
        to cover your living expenses, permanently.
      </div>
    </div>

    <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px;max-width:580px;margin:0 auto 40px">
      <div style="background:#0f0f16;border:0.5px solid rgba(255,255,255,0.07);border-radius:10px;padding:18px 16px;text-align:center">
        <div style="font-size:22px;margin-bottom:8px">⏱</div>
        <div style="font-size:12px;font-weight:600;color:#F2F2F2;margin-bottom:4px">3 minutes</div>
        <div style="font-size:11px;color:#52525B;line-height:1.5">To complete the analysis</div>
      </div>
      <div style="background:#0f0f16;border:0.5px solid rgba(255,255,255,0.07);border-radius:10px;padding:18px 16px;text-align:center">
        <div style="font-size:22px;margin-bottom:8px">📊</div>
        <div style="font-size:12px;font-weight:600;color:#F2F2F2;margin-bottom:4px">3 questions</div>
        <div style="font-size:11px;color:#52525B;line-height:1.5">No complex data needed</div>
      </div>
      <div style="background:#0f0f16;border:0.5px solid rgba(255,255,255,0.07);border-radius:10px;padding:18px 16px;text-align:center">
        <div style="font-size:22px;margin-bottom:8px">📄</div>
        <div style="font-size:12px;font-weight:600;color:#F2F2F2;margin-bottom:4px">Free report</div>
        <div style="font-size:11px;color:#52525B;line-height:1.5">PDF with your results</div>
      </div>
    </div>

    <div style="background:#0f0f16;border:0.5px solid rgba(255,255,255,0.06);border-left:2px solid #5A54C4;
         border-radius:0 10px 10px 0;padding:16px 20px;max-width:580px;margin:0 auto 40px">
      <div style="font-size:13px;color:#A1A1AA;line-height:1.7">
        <strong style="color:#F2F2F2">What you will need:</strong>
        your current age, how much you can invest per month, and an idea of the monthly income
        you would want in retirement. That is all. We handle the math.
      </div>
    </div>
    """, unsafe_allow_html=True)

    col_start = st.columns([2, 1, 2])
    with col_start[1]:
        if st.button("Start my analysis →"):
            st.session_state.step = 1
            st.rerun()

elif st.session_state.step <= 3:
    step_indicator(st.session_state.step)
    st.markdown("<div style='margin-bottom:32px'></div>", unsafe_allow_html=True)

# ── STEP 1 ──
if st.session_state.step == 1:
    st.markdown('<div style="font-family:\'Space Grotesk\',sans-serif;font-size:28px;font-weight:700;letter-spacing:-.5px;margin-bottom:6px">Your current situation</div>', unsafe_allow_html=True)
    st.markdown('<div style="color:#71717A;font-size:15px;margin-bottom:32px">A few numbers about where you are today. No income data needed — just your age, what you invest, and your time horizon.</div>', unsafe_allow_html=True)

    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown(label("Current age"), unsafe_allow_html=True)
        current_age = st.number_input("Age", min_value=18, max_value=70, value=30, step=1, label_visibility="collapsed")
        st.markdown('<div style="font-size:11px;color:#3f3f46;margin-top:4px">Your age today</div>', unsafe_allow_html=True)
    with col2:
        st.markdown(label("Monthly investment (€)"), unsafe_allow_html=True)
        monthly_investment = st.number_input("Monthly investment", min_value=0, max_value=50000, value=300, step=50, label_visibility="collapsed")
        st.markdown('<div style="font-size:11px;color:#3f3f46;margin-top:4px">What you can consistently set aside each month. €200–800 is typical for someone starting out.</div>', unsafe_allow_html=True)
    with col3:
        st.markdown(label("Years of accumulation"), unsafe_allow_html=True)
        years_accumulation = st.slider("Years acc", 5, 40, 25, label_visibility="collapsed")
        st.markdown('<div style="font-size:11px;color:#3f3f46;margin-top:4px">How many years you plan to invest before retiring. Most people choose 20–30 years.</div>', unsafe_allow_html=True)

    ret_age = current_age + years_accumulation
    st.markdown(f"""
    {card(f'''
      <div style="display:flex;justify-content:space-between;align-items:center">
        <div>
          <div style="font-size:13px;color:#52525B;margin-bottom:4px">Estimated retirement age</div>
          <div style="font-size:32px;font-weight:700;color:#7F77DD;font-family:'Space Grotesk',sans-serif">{ret_age}</div>
        </div>
        <div style="text-align:right">
          <div style="font-size:13px;color:#52525B;margin-bottom:4px">Monthly investment</div>
          <div style="font-size:24px;font-weight:700;color:#F2F2F2;font-family:'Space Grotesk',sans-serif">€{monthly_investment:,}</div>
        </div>
      </div>
    ''')}""", unsafe_allow_html=True)

    st.markdown("<div style='margin-top:8px'></div>", unsafe_allow_html=True)
    col_btn = st.columns([2, 1, 2])
    with col_btn[1]:
        if st.button("Continue →"):
            st.session_state.inputs.update({"monthly_investment": monthly_investment, "years_accumulation": years_accumulation, "current_age": current_age})
            st.session_state.step = 2
            st.rerun()

# ── STEP 2 ──
elif st.session_state.step == 2:
    # Summary bar from step 1
    s1 = st.session_state.inputs
    st.markdown(f"""
    <div style="background:#0f0f16;border:0.5px solid rgba(255,255,255,0.06);border-radius:10px;
         padding:12px 20px;margin-bottom:28px;display:flex;gap:32px;flex-wrap:wrap">
      <div><span style="font-size:10px;color:#3f3f46;text-transform:uppercase;letter-spacing:.8px">Age</span>
           <span style="font-size:13px;color:#F2F2F2;margin-left:8px;font-weight:500">{s1.get('current_age',30)}</span></div>
      <div><span style="font-size:10px;color:#3f3f46;text-transform:uppercase;letter-spacing:.8px">Monthly investment</span>
           <span style="font-size:13px;color:#F2F2F2;margin-left:8px;font-weight:500">€{s1.get('monthly_investment',300):,}</span></div>
      <div><span style="font-size:10px;color:#3f3f46;text-transform:uppercase;letter-spacing:.8px">Years accumulation</span>
           <span style="font-size:13px;color:#F2F2F2;margin-left:8px;font-weight:500">{s1.get('years_accumulation',25)}</span></div>
      <div><span style="font-size:10px;color:#3f3f46;text-transform:uppercase;letter-spacing:.8px">Retirement age</span>
           <span style="font-size:13px;color:#7F77DD;margin-left:8px;font-weight:600">{s1.get('current_age',30) + s1.get('years_accumulation',25)}</span></div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div style="font-family:\'Space Grotesk\',sans-serif;font-size:28px;font-weight:700;letter-spacing:-.5px;margin-bottom:6px">Your retirement goal</div>', unsafe_allow_html=True)
    st.markdown('<div style="color:#71717A;font-size:15px;margin-bottom:32px">Define what financial independence looks like for you. Think in terms of your current monthly expenses — not a dream lifestyle, but a life you would genuinely enjoy.</div>', unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        st.markdown(label("Desired monthly income (€)"), unsafe_allow_html=True)
        monthly_income = st.number_input("Monthly income", min_value=500, max_value=20000, value=1500, step=100, label_visibility="collapsed")
        st.markdown('<div style="font-size:11px;color:#3f3f46;margin-top:4px">The monthly income your portfolio needs to generate. Start with your current expenses. €1,000–2,500/month covers a comfortable life in most of Europe.</div>', unsafe_allow_html=True)
    with col2:
        st.markdown(label("Years in retirement"), unsafe_allow_html=True)
        years_retirement = st.slider("Years ret", 5, 40, 25, label_visibility="collapsed")
        st.markdown('<div style="font-size:11px;color:#3f3f46;margin-top:4px">How long your portfolio needs to last. If you retire at 55, plan for 30–35 years. When in doubt, choose more rather than less.</div>', unsafe_allow_html=True)

    fn = freedom_number(monthly_income)
    st.markdown(f"""
    {card(f'''
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:24px">
        <div>
          <div style="font-size:11px;color:#3f3f46;letter-spacing:1px;text-transform:uppercase;margin-bottom:8px">Your Freedom Number</div>
          <div style="font-size:40px;font-weight:700;color:#7F77DD;font-family:'Space Grotesk',sans-serif;line-height:1">€{int(fn):,}</div>
          <div style="font-size:12px;color:#52525B;margin-top:6px">Portfolio needed at 4% withdrawal rate</div>
        </div>
        <div>
          <div style="font-size:11px;color:#3f3f46;letter-spacing:1px;text-transform:uppercase;margin-bottom:8px">Monthly income generated</div>
          <div style="font-size:40px;font-weight:700;color:#F2F2F2;font-family:'Space Grotesk',sans-serif;line-height:1">€{monthly_income:,}</div>
          <div style="font-size:12px;color:#52525B;margin-top:6px">For {years_retirement} years of retirement</div>
        </div>
      </div>
    ''', border_color='rgba(90,84,196,0.25)')}""", unsafe_allow_html=True)

    st.markdown("""
    <div style="background:#0f0f16;border:0.5px solid rgba(255,255,255,0.05);border-left:2px solid #26215C;
         border-radius:0 8px 8px 0;padding:14px 18px;margin-top:4px">
      <div style="font-size:12px;color:#52525B;line-height:1.7">
        <strong style="color:#3f3f46">What is the Freedom Number?</strong>
        It is the total invested portfolio that, at a 4% annual withdrawal rate, generates your desired income indefinitely.
        This is based on the widely-cited 4% rule, validated across multiple market cycles.
        Your number updates as you adjust the sliders above.
      </div>
    </div>
    """, unsafe_allow_html=True)

    c1, c2 = st.columns(2)
    with c1:
        if st.button("← Back"):
            st.session_state.step = 1; st.rerun()
    with c2:
        if st.button("Continue →"):
            st.session_state.inputs.update({"monthly_income": monthly_income, "years_retirement": years_retirement, "freedom_number": fn})
            st.session_state.step = 3; st.rerun()

# ── STEP 3 ──
elif st.session_state.step == 3:
    # Full summary from steps 1 + 2
    s = st.session_state.inputs
    fn_s = freedom_number(s.get('monthly_income', 1500))
    st.markdown(f"""
    <div style="background:#0f0f16;border:0.5px solid rgba(255,255,255,0.06);border-radius:10px;
         padding:12px 20px;margin-bottom:28px;display:flex;gap:28px;flex-wrap:wrap">
      <div><span style="font-size:10px;color:#3f3f46;text-transform:uppercase;letter-spacing:.8px">Age</span>
           <span style="font-size:13px;color:#F2F2F2;margin-left:8px;font-weight:500">{s.get('current_age',30)}</span></div>
      <div><span style="font-size:10px;color:#3f3f46;text-transform:uppercase;letter-spacing:.8px">Monthly investment</span>
           <span style="font-size:13px;color:#F2F2F2;margin-left:8px;font-weight:500">€{s.get('monthly_investment',300):,}</span></div>
      <div><span style="font-size:10px;color:#3f3f46;text-transform:uppercase;letter-spacing:.8px">Accumulation</span>
           <span style="font-size:13px;color:#F2F2F2;margin-left:8px;font-weight:500">{s.get('years_accumulation',25)} yrs</span></div>
      <div><span style="font-size:10px;color:#3f3f46;text-transform:uppercase;letter-spacing:.8px">Monthly income goal</span>
           <span style="font-size:13px;color:#F2F2F2;margin-left:8px;font-weight:500">€{s.get('monthly_income',1500):,}</span></div>
      <div><span style="font-size:10px;color:#3f3f46;text-transform:uppercase;letter-spacing:.8px">Freedom Number</span>
           <span style="font-size:13px;color:#7F77DD;margin-left:8px;font-weight:600">€{int(fn_s):,}</span></div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div style="font-family:\'Space Grotesk\',sans-serif;font-size:28px;font-weight:700;letter-spacing:-.5px;margin-bottom:6px">Your investment profile</div>', unsafe_allow_html=True)
    st.markdown('<div style="color:#71717A;font-size:15px;margin-bottom:32px">Choose how you plan to invest. If you are unsure, <strong style="color:#A1A1AA">Balanced</strong> is the right default for most long-term investors.</div>', unsafe_allow_html=True)

    profile_data = {
        "Conservative": ("~4% annual return", "Low-cost bond and money market funds. Lower growth, lower volatility. Suitable if you are close to retirement or prefer stability over returns.", "#AFA9EC"),
        "Balanced":     ("~6% annual return", "Global equity index funds + bonds. The most recommended long-term allocation for investors with a 15+ year horizon. Based on Vanguard 2024 data.", "#7F77DD"),
        "Aggressive":   ("~8% annual return", "100% global equity index funds. Higher long-term growth, higher short-term volatility. Best suited for investors with 20+ years and strong emotional discipline.", "#5A54C4"),
    }

    profile = st.selectbox("Investment profile", list(profile_data.keys()), index=1, label_visibility="collapsed")
    ret_label, ret_desc, pcolor = profile_data[profile]

    st.markdown(f"""
    {card(f'''
      <div style="display:flex;align-items:center;gap:16px">
        <div style="width:4px;height:60px;background:{pcolor};border-radius:2px;flex-shrink:0"></div>
        <div>
          <div style="font-size:17px;font-weight:600;color:#F2F2F2;margin-bottom:4px">{ret_label}</div>
          <div style="font-size:13px;color:#52525B;line-height:1.65">{ret_desc}</div>
        </div>
      </div>
    ''')}""", unsafe_allow_html=True)

    st.markdown("""
    <div style="background:#0f0f16;border:0.5px solid rgba(255,255,255,0.05);border-left:2px solid #26215C;
         border-radius:0 8px 8px 0;padding:14px 18px;margin-top:4px">
      <div style="font-size:12px;color:#3f3f46;line-height:1.7">
        <strong style="color:#52525B">About these return estimates:</strong>
        Returns are nominal (before inflation) and based on long-term historical averages.
        The S&P 500 has returned ~10% annually since 1928. A globally diversified 60/40 portfolio
        has returned ~6.8% since 1997 (Vanguard, 2024). Past performance does not guarantee future results.
        The Pro version includes inflation-adjusted projections so you can plan in real euros.
      </div>
    </div>
    """, unsafe_allow_html=True)

    c1, c2 = st.columns(2)
    with c1:
        if st.button("← Back"):
            st.session_state.step = 2; st.rerun()
    with c2:
        if st.button("Run simulation →"):
            st.session_state.inputs.update({"profile": profile, "ann_return": get_return(profile)})
            st.session_state.step = 4; st.rerun()

# ─────────────────────────────────────────────
# STEP 4 — RESULTS
# ─────────────────────────────────────────────

elif st.session_state.step == 4:

    inp = st.session_state.inputs
    mi, mc_inc = inp["monthly_investment"], inp["monthly_income"]
    ya, yr     = inp["years_accumulation"],  inp["years_retirement"]
    ar         = inp["ann_return"]
    fn         = inp["freedom_number"]
    age        = inp["current_age"]

    history, survives = simulate(mi, mc_inc, ya, yr, ar)
    inv_needed        = required_investment(mc_inc, ya, yr, ar)
    hist_rec, _       = simulate(inv_needed, mc_inc, ya, yr, ar)
    max_cap           = max(history) if history else 0
    target_cap        = mc_inc * 12 / ar
    score             = min(100, (max_cap / target_cap) * 100)
    ytf               = years_to_fi(mi, fn, ar)
    fi_age            = age + round(ytf) if ytf else None
    gap               = max(0, inv_needed - mi)

    score_color = "#7F77DD" if score >= 70 else "#EF9F27" if score >= 40 else "#E24B4A"
    score_text  = (
        "Strong trajectory." if score >= 80 else
        "Making progress — a few adjustments could help." if score >= 60 else
        "Your plan needs attention." if score >= 40 else
        "Significant gaps detected."
    )

    # ── HEADER ──
    st.markdown(f"""
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:28px">
      <div>
        <div style="font-family:'Space Grotesk',sans-serif;font-size:28px;font-weight:700;letter-spacing:-.5px">Your results</div>
        <div style="font-size:13px;color:#52525B;margin-top:4px">{inp['profile']} profile · {ya} years accumulation · {yr} years retirement</div>
      </div>
      <div style="font-size:12px;color:#3f3f46;cursor:pointer" onclick="">← Start over</div>
    </div>
    {divider('0 0 28px')}
    """, unsafe_allow_html=True)

    # ── SCORE + KEY METRICS ──
    col_score, col_fn, col_gap, col_fi = st.columns(4)

    with col_score:
        st.markdown(f"""
        <div style="background:#0f0f16;border:0.5px solid rgba(255,255,255,0.08);border-radius:14px;padding:24px;text-align:center">
          <div style="font-size:11px;color:#3f3f46;letter-spacing:1px;text-transform:uppercase;margin-bottom:10px">Readiness score</div>
          <div style="font-size:60px;font-weight:700;color:{score_color};font-family:'Space Grotesk',sans-serif;line-height:1">{int(score)}</div>
          <div style="font-size:11px;color:#52525B;margin-top:6px">{score_text}</div>
        </div>
        """, unsafe_allow_html=True)
        st.progress(score / 100)

    with col_fn:
        st.metric("Freedom Number", f"€{int(fn):,}")
    with col_gap:
        if gap == 0:
            st.metric("Monthly gap", "On track ✓", delta="Sufficient savings")
        else:
            st.metric("Monthly gap", f"+€{gap:,}/mo", delta="Needs increase", delta_color="inverse")
    with col_fi:
        st.metric("Estimated FI age", f"Age {fi_age}" if fi_age else "Not reached")

    st.markdown("<div style='margin-top:24px'></div>", unsafe_allow_html=True)

    # ── CHART ──
    x1 = [i / 12 for i in range(len(history))]
    x2 = [i / 12 for i in range(len(hist_rec))]

    fig = go.Figure()
    fig.add_vrect(x0=0, x1=ya, fillcolor="rgba(127,119,221,0.03)", line_width=0, annotation_text="Accumulation", annotation_font_color="#3f3f46", annotation_position="top left")
    fig.add_vrect(x0=ya, x1=max(x1[-1], x2[-1]), fillcolor="rgba(255,255,255,0.01)", line_width=0, annotation_text="Retirement", annotation_font_color="#3f3f46", annotation_position="top left")
    fig.add_trace(go.Scatter(x=x1, y=history,  mode="lines", name="Your plan",        line=dict(color="#7F77DD", width=2.5)))
    fig.add_trace(go.Scatter(x=x2, y=hist_rec, mode="lines", name="Recommended plan", line=dict(color="#f9f09d", width=1.5, dash="dot")))
    fig.add_vline(x=ya, line_color="rgba(255,255,255,0.12)", line_dash="dash")
    fig.update_layout(
        template="plotly_dark", paper_bgcolor="#0B0B0F", plot_bgcolor="#0f0f16",
        height=380, margin=dict(l=0, r=0, t=10, b=0),
        legend=dict(bgcolor="rgba(0,0,0,0)", font=dict(color="#A1A1AA", size=12), orientation="h", y=-0.12),
        xaxis=dict(title="Years", color="#3f3f46", gridcolor="rgba(255,255,255,0.04)", zeroline=False),
        yaxis=dict(title="Portfolio (€)", color="#3f3f46", gridcolor="rgba(255,255,255,0.04)", zeroline=False, tickprefix="€")
    )
    st.plotly_chart(fig, use_container_width=True)

    st.markdown(f"""
    <div style="font-size:11px;color:#3f3f46;line-height:1.7;margin-top:-8px;margin-bottom:8px">
      Projections are shown in nominal terms and do not adjust for inflation.
      Assumed annual return: {int(ar*100)}% ({inp['profile']} profile) · Monte Carlo and inflation-adjusted projections available in Pro.
    </div>
    """, unsafe_allow_html=True)

    # ── INSIGHTS ──
    st.markdown(divider(), unsafe_allow_html=True)
    st.markdown('<div style="font-family:\'Space Grotesk\',sans-serif;font-size:20px;font-weight:700;margin-bottom:20px">What this means for you</div>', unsafe_allow_html=True)

    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown(f"""
        <div style="background:#0f0f16;border:0.5px solid {'rgba(127,119,221,0.3)' if gap==0 else 'rgba(239,159,39,0.3)'};border-radius:12px;padding:20px 18px">
          <div style="font-size:11px;color:#3f3f46;letter-spacing:1px;text-transform:uppercase;margin-bottom:8px">Monthly investment</div>
          <div style="font-size:22px;font-weight:700;font-family:'Space Grotesk',sans-serif;color:{'#7F77DD' if gap==0 else '#EF9F27'};margin-bottom:6px">
            {'On track ✓' if gap==0 else f'+€{gap:,}/month needed'}
          </div>
          <div style="font-size:12px;color:#52525B;line-height:1.6">
            {'Your savings rate is sufficient for your retirement goal.' if gap==0 else f'Increasing by €{gap:,}/month would close the gap completely.'}
          </div>
        </div>
        """, unsafe_allow_html=True)
    with c2:
        dur_text = f"Sustained for full {yr} years" if survives else f"Runs out after ~{(len(history)/12 - ya):.1f} years"
        dur_color = "#7F77DD" if survives else "#E24B4A"
        st.markdown(f"""
        <div style="background:#0f0f16;border:0.5px solid {'rgba(127,119,221,0.3)' if survives else 'rgba(226,75,74,0.3)'};border-radius:12px;padding:20px 18px">
          <div style="font-size:11px;color:#3f3f46;letter-spacing:1px;text-transform:uppercase;margin-bottom:8px">Retirement sustainability</div>
          <div style="font-size:22px;font-weight:700;font-family:'Space Grotesk',sans-serif;color:{dur_color};margin-bottom:6px">{dur_text}</div>
          <div style="font-size:12px;color:#52525B;line-height:1.6">
            {'Your portfolio sustains €' + f'{mc_inc:,}/month for your full retirement.' if survives else 'Consider increasing contributions or reducing desired income.'}
          </div>
        </div>
        """, unsafe_allow_html=True)
    with c3:
        st.markdown(f"""
        <div style="background:#0f0f16;border:0.5px solid rgba(127,119,221,0.2);border-radius:12px;padding:20px 18px">
          <div style="font-size:11px;color:#3f3f46;letter-spacing:1px;text-transform:uppercase;margin-bottom:8px">Financial independence</div>
          <div style="font-size:22px;font-weight:700;font-family:'Space Grotesk',sans-serif;color:#7F77DD;margin-bottom:6px">
            {'Age ' + str(fi_age) if fi_age else 'Not reached'}
          </div>
          <div style="font-size:12px;color:#52525B;line-height:1.6">
            {'At your current savings rate, you could stop depending on a salary at age ' + str(fi_age) + '.' if fi_age else 'Increase your monthly investment to reach FI within your lifetime.'}
          </div>
        </div>
        """, unsafe_allow_html=True)

    # ── PDF GATE ──
    st.markdown(divider("40px 0"), unsafe_allow_html=True)

    if not st.session_state.pdf_unlocked:

        st.markdown(f"""
        <div style="font-family:'Space Grotesk',sans-serif;font-size:22px;font-weight:700;margin-bottom:6px">Your personalised report</div>
        <div style="font-size:14px;color:#71717A;margin-bottom:28px">A full PDF with your projections, score, metrics and insights. Free to download.</div>
        """, unsafe_allow_html=True)

        # ── REPORT PREVIEW (blurred) ──
        st.markdown(f"""
        <div style="position:relative;border-radius:14px;overflow:hidden;margin-bottom:28px">

          <!-- blurred overlay -->
          <div style="position:absolute;inset:0;background:linear-gradient(to bottom, rgba(11,11,15,0.1) 0%, rgba(11,11,15,0.85) 60%, rgba(11,11,15,0.97) 100%);
               backdrop-filter:blur(2px);z-index:2;border-radius:14px;display:flex;align-items:flex-end;justify-content:center;padding-bottom:32px">
            <div style="text-align:center">
              <div style="display:inline-block;background:rgba(90,84,196,0.15);border:0.5px solid rgba(90,84,196,0.4);border-radius:20px;
                color:#9490e8;font-size:11px;letter-spacing:.8px;text-transform:uppercase;padding:5px 14px;margin-bottom:10px">Free report</div>
              <div style="font-family:'Space Grotesk',sans-serif;font-size:16px;font-weight:600;color:#F2F2F2;margin-bottom:4px">Register to unlock your report</div>
              <div style="font-size:12px;color:#52525B">Takes 10 seconds. No payment required.</div>
            </div>
          </div>

          <!-- preview content -->
          <div style="background:#0f0f16;border:0.5px solid rgba(255,255,255,0.07);border-radius:14px;padding:32px;filter:blur(1px)">
            <div style="font-size:10px;color:#3f3f46;letter-spacing:1.2px;text-transform:uppercase;margin-bottom:4px">THE FREEDOM PROJECT (fi)</div>
            <div style="font-family:'Space Grotesk',sans-serif;font-size:22px;font-weight:700;color:#F2F2F2;margin-bottom:20px">Financial Independence Report</div>
            <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:20px">
              <div style="background:#141420;border-radius:8px;padding:14px">
                <div style="font-size:10px;color:#3f3f46;margin-bottom:6px">Readiness score</div>
                <div style="font-size:28px;font-weight:700;color:{score_color};font-family:'Space Grotesk',sans-serif">{int(score)}</div>
              </div>
              <div style="background:#141420;border-radius:8px;padding:14px">
                <div style="font-size:10px;color:#3f3f46;margin-bottom:6px">Freedom Number</div>
                <div style="font-size:18px;font-weight:700;color:#F2F2F2;font-family:'Space Grotesk',sans-serif">€{int(fn):,}</div>
              </div>
              <div style="background:#141420;border-radius:8px;padding:14px">
                <div style="font-size:10px;color:#3f3f46;margin-bottom:6px">Monthly gap</div>
                <div style="font-size:18px;font-weight:700;color:#F2F2F2;font-family:'Space Grotesk',sans-serif">{'✓' if gap==0 else f'+€{gap:,}'}</div>
              </div>
              <div style="background:#141420;border-radius:8px;padding:14px">
                <div style="font-size:10px;color:#3f3f46;margin-bottom:6px">FI age</div>
                <div style="font-size:18px;font-weight:700;color:#F2F2F2;font-family:'Space Grotesk',sans-serif">{fi_age if fi_age else '—'}</div>
              </div>
            </div>
            <div style="height:48px;background:#141420;border-radius:8px;margin-bottom:10px"></div>
            <div style="height:32px;background:#141420;border-radius:8px;width:70%"></div>
          </div>
        </div>
        """, unsafe_allow_html=True)

        # ── REGISTRATION FORM ──
        st.markdown(f"""
        <div style="background:rgba(90,84,196,0.06);border:0.5px solid rgba(90,84,196,0.2);border-radius:14px;padding:28px 32px">
          <div style="font-family:'Space Grotesk',sans-serif;font-size:18px;font-weight:700;margin-bottom:4px">Get your free report</div>
          <div style="font-size:13px;color:#52525B;margin-bottom:24px">Enter your details to download your personalised financial independence report.</div>
        """, unsafe_allow_html=True)

        fc1, fc2, fc3 = st.columns(3)
        with fc1:
            st.markdown('<div style="font-size:11px;color:#3f3f46;letter-spacing:1px;text-transform:uppercase;margin-bottom:6px">Full name</div>', unsafe_allow_html=True)
            name = st.text_input("name", placeholder="Arturo Maldonado", label_visibility="collapsed")
        with fc2:
            st.markdown('<div style="font-size:11px;color:#3f3f46;letter-spacing:1px;text-transform:uppercase;margin-bottom:6px">Email address</div>', unsafe_allow_html=True)
            email = st.text_input("email", placeholder="you@example.com", label_visibility="collapsed")
        with fc3:
            st.markdown('<div style="font-size:11px;color:#3f3f46;letter-spacing:1px;text-transform:uppercase;margin-bottom:6px">Country</div>', unsafe_allow_html=True)
            country = st.selectbox("country", [
                "Germany", "Netherlands", "Switzerland", "Austria", "Belgium",
                "Sweden", "Denmark", "Norway", "Finland", "Luxembourg",
                "Spain", "Portugal", "France", "Italy", "United Kingdom",
                "Ireland", "Poland", "Czech Republic", "Estonia", "Latvia",
                "United States", "Canada", "Australia", "New Zealand",
                "Mexico", "Colombia", "Argentina", "Chile", "Peru",
                "Brazil", "Costa Rica", "Uruguay",
                "Singapore", "Hong Kong", "Japan", "South Korea",
                "United Arab Emirates", "South Africa", "Israel",
                "Other"
            ], label_visibility="collapsed")

        st.markdown("</div>", unsafe_allow_html=True)
        st.markdown("<div style='margin-top:16px'></div>", unsafe_allow_html=True)

        col_dl = st.columns([2, 1, 2])
        with col_dl[1]:
            if st.button("Download my report →"):
                email_valid = re.match(r'^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$', email)
                if not name.strip():
                    st.error("Please enter your name.")
                elif not email.strip() or not email_valid:
                    st.error("Please enter a valid email address.")
                else:
                    st.session_state.inputs.update({"name": name, "email": email, "country": country})
                    save_lead("free_leads", {
                        "Email":          email,
                        "Country":        country,
                        "Score":          round(score, 1),
                        "Freedom Number": int(fn),
                        "Profile":        inp.get("profile", ""),
                    })
                    st.session_state.pdf_unlocked = True
                    st.rerun()

    else:
        # ── PDF READY ──
        inp2 = st.session_state.inputs
        first_name = inp2.get("name", "").split()[0] if inp2.get("name") else "there"

        st.markdown(f"""
        <div style="background:rgba(90,84,196,0.07);border:0.5px solid rgba(90,84,196,0.25);border-radius:14px;padding:28px 32px;
             display:flex;align-items:center;gap:24px;margin-bottom:20px">
          <div style="width:48px;height:48px;border-radius:50%;background:rgba(90,84,196,0.2);border:1px solid rgba(90,84,196,0.4);
               display:flex;align-items:center;justify-content:center;font-size:20px;flex-shrink:0">✓</div>
          <div>
            <div style="font-family:'Space Grotesk',sans-serif;font-size:18px;font-weight:700;margin-bottom:4px">
              Your report is ready, {first_name}.
            </div>
            <div style="font-size:13px;color:#52525B">
              Your personalised Financial Independence Report is ready to download.
              Prepared for {inp2.get('email','')} · {inp2.get('country','')}.
            </div>
          </div>
        </div>
        """, unsafe_allow_html=True)

        pdf = generate_pdf(inp2, score, inv_needed, max_cap, target_cap, fn, fi_age, survives, gap)
        col_dl2 = st.columns([2, 1, 2])
        with col_dl2[1]:
            st.download_button(
                label="⬇ Download PDF report",
                data=pdf,
                file_name=f"freedom_project_report_{datetime.now().strftime('%Y%m%d')}.pdf",
                mime="application/pdf"
            )

    # ── PRO SECTION ──
    st.markdown(divider("32px 0"), unsafe_allow_html=True)

    pro = st.session_state.pro_unlocked

    if not pro:
        # ── PRO GATE ──
        st.markdown(f"""
        <div style="background:#0f0f16;border:0.5px solid rgba(90,84,196,0.25);border-radius:14px;overflow:hidden">

          <!-- header -->
          <div style="background:rgba(90,84,196,0.1);padding:24px 28px;border-bottom:0.5px solid rgba(90,84,196,0.2)">
            <div style="display:inline-block;background:rgba(90,84,196,0.2);border:0.5px solid rgba(90,84,196,0.4);
                 border-radius:20px;color:#9490e8;font-size:10px;letter-spacing:.8px;text-transform:uppercase;
                 padding:4px 12px;margin-bottom:12px">Pro access</div>
            <div style="font-family:'Space Grotesk',sans-serif;font-size:22px;font-weight:700;margin-bottom:6px">
              Go deeper with your analysis
            </div>
            <div style="font-size:14px;color:#71717A;line-height:1.7;max-width:520px">
              The Pro version gives you a more complete picture of your financial future,
              including probabilistic analysis, real returns and a detailed branded report.
            </div>
          </div>

          <!-- features grid -->
          <div style="padding:24px 28px">
            <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px;margin-bottom:20px">
              <div style="background:#141420;border:0.5px solid rgba(255,255,255,0.06);border-radius:9px;padding:16px">
                <div style="font-size:16px;margin-bottom:8px">📊</div>
                <div style="font-size:13px;font-weight:600;color:#F2F2F2;margin-bottom:4px">Monte Carlo</div>
                <div style="font-size:12px;color:#52525B;line-height:1.5">1,000 simulations. Know the real probability your plan survives.</div>
              </div>
              <div style="background:#141420;border:0.5px solid rgba(255,255,255,0.06);border-radius:9px;padding:16px">
                <div style="font-size:16px;margin-bottom:8px">📉</div>
                <div style="font-size:13px;font-weight:600;color:#F2F2F2;margin-bottom:4px">Inflation-adjusted</div>
                <div style="font-size:12px;color:#52525B;line-height:1.5">See your projections in real euros, not just nominal figures.</div>
              </div>
              <div style="background:#141420;border:0.5px solid rgba(255,255,255,0.06);border-radius:9px;padding:16px">
                <div style="font-size:16px;margin-bottom:8px">🔀</div>
                <div style="font-size:13px;font-weight:600;color:#F2F2F2;margin-bottom:4px">Scenario comparison</div>
                <div style="font-size:12px;color:#52525B;line-height:1.5">Conservative vs balanced vs aggressive — side by side.</div>
              </div>
              <div style="background:#141420;border:0.5px solid rgba(255,255,255,0.06);border-radius:9px;padding:16px">
                <div style="font-size:16px;margin-bottom:8px">📄</div>
                <div style="font-size:13px;font-weight:600;color:#F2F2F2;margin-bottom:4px">Full PDF report</div>
                <div style="font-size:12px;color:#52525B;line-height:1.5">4-page branded report with chart, metrics and personalised insights.</div>
              </div>
              <div style="background:#141420;border:0.5px solid rgba(255,255,255,0.06);border-radius:9px;padding:16px">
                <div style="font-size:16px;margin-bottom:8px">📅</div>
                <div style="font-size:13px;font-weight:600;color:#F2F2F2;margin-bottom:4px">FI date estimate</div>
                <div style="font-size:12px;color:#52525B;line-height:1.5">The exact age at which your investments work harder than you do.</div>
              </div>
              <div style="background:#141420;border:0.5px solid rgba(255,255,255,0.06);border-radius:9px;padding:16px">
                <div style="font-size:16px;margin-bottom:8px">🔜</div>
                <div style="font-size:13px;font-weight:600;color:#F2F2F2;margin-bottom:4px">More tools coming</div>
                <div style="font-size:12px;color:#52525B;line-height:1.5">Retirement Planner and Portfolio Allocator. Pro members get early access.</div>
              </div>
            </div>
          </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("""
        <div style="background:rgba(90,84,196,0.07);border:0.5px solid rgba(90,84,196,0.2);border-radius:9px;
             padding:14px 18px;margin-bottom:20px;font-size:12px;color:#71717A;line-height:1.6">
          <span style="color:#9490e8;font-weight:500">Early access:</span>
          Pro is currently available by invitation. Enter your details and access code to unlock.
        </div>
        """, unsafe_allow_html=True)

        # Pro gate form — step 1: email + name
        if not st.session_state.pro_gate_open:
            st.markdown("<div style='margin-top:16px'></div>", unsafe_allow_html=True)
            col_pe1, col_pe2, col_pe3 = st.columns([2, 2, 1])
            with col_pe1:
                st.markdown('<div style="font-size:11px;color:#3f3f46;letter-spacing:1px;text-transform:uppercase;margin-bottom:6px">Email address</div>', unsafe_allow_html=True)
                pro_email = st.text_input("pro_email", placeholder="you@example.com", label_visibility="collapsed")
            with col_pe2:
                st.markdown('<div style="font-size:11px;color:#3f3f46;letter-spacing:1px;text-transform:uppercase;margin-bottom:6px">Full name</div>', unsafe_allow_html=True)
                pro_name = st.text_input("pro_name", placeholder="Arturo Maldonado", label_visibility="collapsed")
            with col_pe3:
                st.markdown('<div style="font-size:11px;color:#3f3f46;margin-bottom:6px">&nbsp;</div>', unsafe_allow_html=True)
                if st.button("Continue →"):
                    pro_email_valid = re.match(r'^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$', pro_email or "")
                    if not pro_name.strip():
                        st.error("Please enter your name.")
                    elif not pro_email_valid:
                        st.error("Please enter a valid email address.")
                    else:
                        st.session_state.inputs.update({"pro_name": pro_name, "pro_email": pro_email})
                        st.session_state.pro_gate_open = True
                        st.rerun()

        # Step 2: access code
        else:
            pro_name  = st.session_state.inputs.get("pro_name", "")
            pro_email = st.session_state.inputs.get("pro_email", "")

            st.markdown(f"""
            <div style="background:#0f0f16;border:0.5px solid rgba(255,255,255,0.07);border-radius:10px;
                 padding:16px 20px;margin-bottom:16px;font-size:13px;color:#71717A">
              Registered as <strong style="color:#F2F2F2">{pro_name}</strong> · {pro_email}
            </div>
            """, unsafe_allow_html=True)

            col_code1, col_code2 = st.columns([3, 1])
            with col_code1:
                st.markdown('<div style="font-size:11px;color:#3f3f46;letter-spacing:1px;text-transform:uppercase;margin-bottom:6px">Access code</div>', unsafe_allow_html=True)
                access_code = st.text_input("access_code", placeholder="Enter your Pro access code", label_visibility="collapsed")
            with col_code2:
                st.markdown('<div style="font-size:11px;color:#3f3f46;margin-bottom:6px">&nbsp;</div>', unsafe_allow_html=True)
                if st.button("Unlock Pro →"):
                    PRO_CODE = st.secrets.get("PRO_ACCESS_CODE", "FREEDOM2025")
                    if access_code.strip().upper() == PRO_CODE.upper():
                        save_lead("pro_leads", {
                            "Name":  pro_name,
                            "Email": pro_email,
                        })
                        st.session_state.pro_unlocked = True
                        st.session_state.pro_email = pro_email
                        st.rerun()
                    else:
                        st.error("Invalid access code. Contact us at thefreedomprojectfi@gmail.com to get yours.")

            if st.button("← Back"):
                st.session_state.pro_gate_open = False
                st.rerun()

    else:
        # ── PRO FEATURES UNLOCKED ──
        st.markdown(f"""
        <div style="display:flex;align-items:center;gap:12px;margin-bottom:24px">
          <div style="display:inline-block;background:rgba(90,84,196,0.15);border:0.5px solid rgba(90,84,196,0.35);
               border-radius:20px;color:#9490e8;font-size:10px;letter-spacing:.8px;text-transform:uppercase;padding:4px 12px">
            Pro access unlocked
          </div>
          <div style="font-size:12px;color:#3f3f46">{st.session_state.pro_email}</div>
        </div>
        """, unsafe_allow_html=True)

        # ── QUICK EDIT PANEL ──
        st.markdown("""
        <div style="background:#0f0f16;border:0.5px solid rgba(90,84,196,0.2);border-radius:14px;padding:24px 28px;margin-bottom:28px">
          <div style="font-family:'Space Grotesk',sans-serif;font-size:15px;font-weight:700;margin-bottom:4px">⚡ Quick edit</div>
          <div style="font-size:12px;color:#52525B;margin-bottom:20px">Adjust your inputs and recalculate instantly without going back through the wizard.</div>
        </div>
        """, unsafe_allow_html=True)

        qe1, qe2, qe3, qe4, qe5, qe6 = st.columns(6)
        with qe1:
            st.markdown('<div style="font-size:10px;color:#3f3f46;letter-spacing:1px;text-transform:uppercase;margin-bottom:6px">Monthly invest (€)</div>', unsafe_allow_html=True)
            new_mi = st.number_input("qe_mi", min_value=0, max_value=50000, value=int(mi), step=50, label_visibility="collapsed")
        with qe2:
            st.markdown('<div style="font-size:10px;color:#3f3f46;letter-spacing:1px;text-transform:uppercase;margin-bottom:6px">Retirement income (€)</div>', unsafe_allow_html=True)
            new_mc = st.number_input("qe_mc", min_value=500, max_value=20000, value=int(mc_inc), step=100, label_visibility="collapsed")
        with qe3:
            st.markdown('<div style="font-size:10px;color:#3f3f46;letter-spacing:1px;text-transform:uppercase;margin-bottom:6px">Years accumulation</div>', unsafe_allow_html=True)
            new_ya = st.number_input("qe_ya", min_value=5, max_value=40, value=int(ya), step=1, label_visibility="collapsed")
        with qe4:
            st.markdown('<div style="font-size:10px;color:#3f3f46;letter-spacing:1px;text-transform:uppercase;margin-bottom:6px">Years retirement</div>', unsafe_allow_html=True)
            new_yr = st.number_input("qe_yr", min_value=5, max_value=40, value=int(yr), step=1, label_visibility="collapsed")
        with qe5:
            st.markdown('<div style="font-size:10px;color:#3f3f46;letter-spacing:1px;text-transform:uppercase;margin-bottom:6px">Profile</div>', unsafe_allow_html=True)
            new_profile = st.selectbox("qe_profile", ["Conservative", "Balanced", "Aggressive", "Custom"],
                                       index=["Conservative","Balanced","Aggressive"].index(inp.get("profile","Balanced"))
                                       if inp.get("profile","Balanced") in ["Conservative","Balanced","Aggressive"] else 3,
                                       label_visibility="collapsed")
        with qe6:
            st.markdown('<div style="font-size:10px;color:#3f3f46;letter-spacing:1px;text-transform:uppercase;margin-bottom:6px">Annual return (%)</div>', unsafe_allow_html=True)
            default_ar = int(ar * 100) if new_profile == "Custom" else {"Conservative":4,"Balanced":6,"Aggressive":8}.get(new_profile, 6)
            new_ar_pct = st.number_input("qe_ar", min_value=1, max_value=20, value=default_ar, step=1,
                                          label_visibility="collapsed",
                                          disabled=(new_profile != "Custom"))
            if new_profile != "Custom":
                st.markdown(f'<div style="font-size:10px;color:#3f3f46;margin-top:3px">{default_ar}% (fixed)</div>', unsafe_allow_html=True)

        final_ar = (new_ar_pct / 100) if new_profile == "Custom" else {"Conservative":0.04,"Balanced":0.06,"Aggressive":0.08}[new_profile]

        col_qe_btn1, col_qe_btn2 = st.columns([1, 4])
        with col_qe_btn1:
            if st.button("↻ Recalculate"):
                st.session_state.inputs.update({
                    "monthly_investment": new_mi,
                    "monthly_income":     new_mc,
                    "years_accumulation": new_ya,
                    "years_retirement":   new_yr,
                    "profile":            new_profile,
                    "ann_return":         final_ar,
                    "freedom_number":     freedom_number(new_mc),
                })
                st.rerun()

        # ── SCENARIOS ──
        if "saved_scenarios" not in st.session_state:
            st.session_state.saved_scenarios = []

        # Save scenario button
        scenario_name = st.text_input("scenario_name", placeholder="Name this scenario (e.g. 'Aggressive plan')",
                                       label_visibility="collapsed")
        col_sv1, col_sv2 = st.columns([1, 5])
        with col_sv1:
            if st.button("💾 Save scenario"):
                if scenario_name.strip():
                    new_scenario = {
                        "name":    scenario_name.strip(),
                        "mi":      mi, "mc":    mc_inc, "ya":  ya,
                        "yr":      yr, "profile": inp.get("profile","Balanced"),
                        "score":   int(score), "fn":    int(fn),
                        "fi_age":  fi_age, "ar":   ar,
                    }
                    # avoid duplicates
                    st.session_state.saved_scenarios = [
                        s for s in st.session_state.saved_scenarios if s["name"] != scenario_name.strip()
                    ]
                    st.session_state.saved_scenarios.append(new_scenario)
                    st.success(f"Scenario '{scenario_name}' saved.")

        # Saved scenarios table
        if st.session_state.saved_scenarios:
            st.markdown(divider("16px 0"), unsafe_allow_html=True)
            st.markdown('<div style="font-family:\'Space Grotesk\',sans-serif;font-size:15px;font-weight:700;margin-bottom:12px">Saved scenarios</div>', unsafe_allow_html=True)

            rows_html = ""
            for s in st.session_state.saved_scenarios:
                score_c = "#7F77DD" if s["score"] >= 70 else "#EF9F27" if s["score"] >= 40 else "#E24B4A"
                rows_html += f"""
                <tr style="border-bottom:0.5px solid rgba(255,255,255,0.05)">
                  <td style="padding:10px 12px;font-size:13px;color:#F2F2F2;font-weight:500">{s['name']}</td>
                  <td style="padding:10px 12px;font-size:13px;color:#52525B">€{s['mi']:,}/mo · {int(s['ar']*100)}%</td>
                  <td style="padding:10px 12px;font-size:13px;color:#F2F2F2">€{s['fn']:,}</td>
                  <td style="padding:10px 12px;font-size:13px;color:{score_c};font-weight:600">{s['score']}/100</td>
                  <td style="padding:10px 12px;font-size:13px;color:#71717A">{'Age '+str(s['fi_age']) if s['fi_age'] else '—'}</td>
                </tr>"""

            st.markdown(f"""
            <table style="width:100%;border-collapse:collapse;background:#0f0f16;border-radius:10px;overflow:hidden;border:0.5px solid rgba(255,255,255,0.07)">
              <thead>
                <tr style="background:#141420;border-bottom:0.5px solid rgba(255,255,255,0.08)">
                  <th style="padding:10px 12px;font-size:10px;color:#3f3f46;letter-spacing:1px;text-transform:uppercase;text-align:left;font-weight:500">Scenario</th>
                  <th style="padding:10px 12px;font-size:10px;color:#3f3f46;letter-spacing:1px;text-transform:uppercase;text-align:left;font-weight:500">Inputs</th>
                  <th style="padding:10px 12px;font-size:10px;color:#3f3f46;letter-spacing:1px;text-transform:uppercase;text-align:left;font-weight:500">Freedom Number</th>
                  <th style="padding:10px 12px;font-size:10px;color:#3f3f46;letter-spacing:1px;text-transform:uppercase;text-align:left;font-weight:500">Score</th>
                  <th style="padding:10px 12px;font-size:10px;color:#3f3f46;letter-spacing:1px;text-transform:uppercase;text-align:left;font-weight:500">FI Age</th>
                </tr>
              </thead>
              <tbody>{rows_html}</tbody>
            </table>
            """, unsafe_allow_html=True)

            # Load scenario buttons
            st.markdown("<div style='margin-top:10px;display:flex;gap:8px;flex-wrap:wrap'>", unsafe_allow_html=True)
            cols_load = st.columns(min(len(st.session_state.saved_scenarios), 5))
            for i, s in enumerate(st.session_state.saved_scenarios):
                with cols_load[i % len(cols_load)]:
                    if st.button(f"Load '{s['name']}'", key=f"load_{i}"):
                        st.session_state.inputs.update({
                            "monthly_investment": s["mi"],
                            "monthly_income":     s["mc"],
                            "years_accumulation": s["ya"],
                            "years_retirement":   s["yr"],
                            "profile":            s["profile"],
                            "ann_return":         s["ar"],
                            "freedom_number":     freedom_number(s["mc"]),
                        })
                        st.rerun()

            if st.button("🗑 Clear all scenarios"):
                st.session_state.saved_scenarios = []
                st.rerun()

        st.markdown(divider("24px 0"), unsafe_allow_html=True)

        # Monte Carlo
        st.markdown('<div style="font-family:\'Space Grotesk\',sans-serif;font-size:18px;font-weight:700;margin-bottom:4px">📊 Monte Carlo simulation</div>', unsafe_allow_html=True)
        st.markdown("<div style='font-size:13px;color:#71717A;margin-bottom:14px'>1,000 simulations with randomised annual returns to estimate the real probability of your plan succeeding.</div>", unsafe_allow_html=True)

        with st.spinner("Running 1,000 simulations..."):
            mc_prob = monte_carlo(mi, mc_inc, ya, yr, ar)
        mc_color = "#7F77DD" if mc_prob >= 70 else "#EF9F27" if mc_prob >= 50 else "#E24B4A"
        mc_label = "Strong probability of success." if mc_prob >= 70 else "Moderate. Consider increasing contributions." if mc_prob >= 50 else "Low. Plan adjustments recommended."

        st.markdown(f"""
        <div style="background:#0f0f16;border:0.5px solid rgba(255,255,255,0.08);border-radius:12px;
             padding:24px 28px;display:flex;align-items:center;gap:24px;margin-bottom:28px">
          <div style="font-size:56px;font-weight:700;color:{mc_color};font-family:'Space Grotesk',sans-serif;line-height:1;flex-shrink:0">{mc_prob}%</div>
          <div>
            <div style="font-size:15px;font-weight:500;color:#F2F2F2;margin-bottom:4px">Probability of success</div>
            <div style="font-size:13px;color:#52525B;margin-bottom:6px">Your plan survives all {yr} years of retirement in {mc_prob}% of simulated market scenarios.</div>
            <div style="font-size:12px;color:{mc_color}">{mc_label}</div>
          </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown(divider("0 0 20px"), unsafe_allow_html=True)

        # Inflation-adjusted
        st.markdown('<div style="font-family:\'Space Grotesk\',sans-serif;font-size:18px;font-weight:700;margin-bottom:4px">📉 Inflation-adjusted projection</div>', unsafe_allow_html=True)
        st.markdown("<div style='font-size:13px;color:#71717A;margin-bottom:14px'>See what your portfolio looks like in real euros after accounting for inflation.</div>", unsafe_allow_html=True)

        inflation = st.slider("Annual inflation rate", 1.0, 5.0, 2.5, 0.5, format="%.1f%%")
        real_return = max(ar - inflation / 100, 0.01)
        hist_real, _ = simulate(mi, mc_inc, ya, yr, real_return)

        fig2 = go.Figure()
        fig2.add_trace(go.Scatter(x=x1, y=history, mode="lines", name="Nominal", line=dict(color="#7F77DD", width=2)))
        fig2.add_trace(go.Scatter(x=[i/12 for i in range(len(hist_real))], y=hist_real, mode="lines",
                                  name=f"Real ({inflation}% inflation)", line=dict(color="#f9f09d", width=2, dash="dot")))
        fig2.update_layout(template="plotly_dark", paper_bgcolor="#0B0B0F", plot_bgcolor="#0f0f16",
                           height=300, margin=dict(l=0,r=0,t=10,b=0),
                           legend=dict(bgcolor="rgba(0,0,0,0)", font=dict(color="#A1A1AA")),
                           xaxis=dict(color="#3f3f46", gridcolor="rgba(255,255,255,0.04)"),
                           yaxis=dict(color="#3f3f46", gridcolor="rgba(255,255,255,0.04)", tickprefix="€"))
        st.plotly_chart(fig2, use_container_width=True)

        st.markdown(divider("0 0 20px"), unsafe_allow_html=True)

        # Scenario comparison
        st.markdown('<div style="font-family:\'Space Grotesk\',sans-serif;font-size:18px;font-weight:700;margin-bottom:4px">🔀 Scenario comparison</div>', unsafe_allow_html=True)
        st.markdown("<div style='font-size:13px;color:#71717A;margin-bottom:14px'>Conservative, balanced and aggressive profiles — your plan across all three.</div>", unsafe_allow_html=True)

        fig3 = go.Figure()
        for p, r, c in [("Conservative", 0.04, "#AFA9EC"), ("Balanced", 0.06, "#7F77DD"), ("Aggressive", 0.08, "#26215C")]:
            h, _ = simulate(mi, mc_inc, ya, yr, r)
            fig3.add_trace(go.Scatter(x=[i/12 for i in range(len(h))], y=h, mode="lines",
                                      name=f"{p} ({int(r*100)}%)", line=dict(color=c, width=2)))
        fig3.update_layout(template="plotly_dark", paper_bgcolor="#0B0B0F", plot_bgcolor="#0f0f16",
                           height=300, margin=dict(l=0,r=0,t=10,b=0),
                           legend=dict(bgcolor="rgba(0,0,0,0)", font=dict(color="#A1A1AA")),
                           xaxis=dict(color="#3f3f46", gridcolor="rgba(255,255,255,0.04)"),
                           yaxis=dict(color="#3f3f46", gridcolor="rgba(255,255,255,0.04)", tickprefix="€"))
        st.plotly_chart(fig3, use_container_width=True)

        st.markdown(divider("0 0 20px"), unsafe_allow_html=True)

        # ── BREAKDOWN: CONTRIBUTIONS VS RETURNS ──
        st.markdown('<div style="font-family:\'Space Grotesk\',sans-serif;font-size:18px;font-weight:700;margin-bottom:4px">📦 Contributions vs returns breakdown</div>', unsafe_allow_html=True)
        st.markdown("<div style='font-size:13px;color:#71717A;margin-bottom:14px'>How much of your final portfolio comes from what you invested versus what the market generated.</div>", unsafe_allow_html=True)

        years_range   = list(range(1, ya + 1))
        contrib_vals  = [mi * 12 * y for y in years_range]
        portfolio_vals = []
        cap_tmp = 0
        mr_tmp  = ar / 12
        for y in years_range:
            for _ in range(12):
                cap_tmp = cap_tmp * (1 + mr_tmp) + mi
            portfolio_vals.append(cap_tmp)
        returns_vals = [max(0, p - c) for p, c in zip(portfolio_vals, contrib_vals)]

        fig_break = go.Figure()
        fig_break.add_trace(go.Bar(x=years_range, y=contrib_vals, name="Your contributions",
                                   marker_color="#26215C"))
        fig_break.add_trace(go.Bar(x=years_range, y=returns_vals, name="Market returns",
                                   marker_color="#7F77DD"))
        fig_break.update_layout(
            barmode="stack", template="plotly_dark",
            paper_bgcolor="#0B0B0F", plot_bgcolor="#0f0f16",
            height=320, margin=dict(l=0,r=0,t=10,b=0),
            legend=dict(bgcolor="rgba(0,0,0,0)", font=dict(color="#A1A1AA")),
            xaxis=dict(title="Year", color="#3f3f46", gridcolor="rgba(255,255,255,0.04)"),
            yaxis=dict(title="Portfolio (€)", color="#3f3f46", gridcolor="rgba(255,255,255,0.04)", tickprefix="€")
        )
        st.plotly_chart(fig_break, use_container_width=True)

        final_contrib_pct = round((contrib_vals[-1] / max(portfolio_vals[-1], 1)) * 100)
        final_returns_pct = 100 - final_contrib_pct
        st.markdown(f"""
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:8px">
          <div style="background:#0f0f16;border:0.5px solid rgba(255,255,255,0.07);border-radius:10px;padding:16px 20px">
            <div style="font-size:11px;color:#3f3f46;margin-bottom:4px">YOUR CONTRIBUTIONS</div>
            <div style="font-size:24px;font-weight:700;color:#AFA9EC;font-family:'Space Grotesk',sans-serif">€{int(contrib_vals[-1]):,}</div>
            <div style="font-size:12px;color:#52525B">{final_contrib_pct}% of final portfolio</div>
          </div>
          <div style="background:#0f0f16;border:0.5px solid rgba(255,255,255,0.07);border-radius:10px;padding:16px 20px">
            <div style="font-size:11px;color:#3f3f46;margin-bottom:4px">MARKET RETURNS</div>
            <div style="font-size:24px;font-weight:700;color:#7F77DD;font-family:'Space Grotesk',sans-serif">€{int(returns_vals[-1]):,}</div>
            <div style="font-size:12px;color:#52525B">{final_returns_pct}% of final portfolio</div>
          </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown(divider("24px 0"), unsafe_allow_html=True)

        # ── COAST FI ──
        st.markdown('<div style="font-family:\'Space Grotesk\',sans-serif;font-size:18px;font-weight:700;margin-bottom:4px">🏄 Coast FI calculator</div>', unsafe_allow_html=True)
        st.markdown("<div style='font-size:13px;color:#71717A;margin-bottom:14px'>The point at which you can stop investing and compounding alone will grow your portfolio to your Freedom Number by retirement age.</div>", unsafe_allow_html=True)

        def coast_fi_number(fn_target, years_to_retirement, ann_return):
            return fn_target / ((1 + ann_return) ** years_to_retirement)

        ret_age = age + ya
        coast_fi_vals = []
        coast_years   = list(range(1, ya))
        for y in coast_years:
            yrs_left = ret_age - (age + y)
            if yrs_left > 0:
                coast_fi_vals.append(coast_fi_number(fn, yrs_left, ar))
            else:
                coast_fi_vals.append(fn)

        coast_reached_age = None
        cap_coast = 0
        for m in range(ya * 12):
            cap_coast = cap_coast * (1 + ar/12) + mi
            yr_now = age + m/12
            yrs_left = ret_age - yr_now
            if yrs_left > 0:
                coast_needed = coast_fi_number(fn, yrs_left, ar)
                if cap_coast >= coast_needed and coast_reached_age is None:
                    coast_reached_age = round(yr_now, 1)

        fig_coast = go.Figure()
        cap_by_year = []
        c = 0
        for y in coast_years:
            for _ in range(12): c = c * (1 + ar/12) + mi
            cap_by_year.append(c)

        fig_coast.add_trace(go.Scatter(x=[age + y for y in coast_years], y=coast_fi_vals,
                                       mode="lines", name="Coast FI target",
                                       line=dict(color="#f9f09d", width=2, dash="dot")))
        fig_coast.add_trace(go.Scatter(x=[age + y for y in coast_years], y=cap_by_year,
                                       mode="lines", name="Your portfolio",
                                       line=dict(color="#7F77DD", width=2)))
        fig_coast.update_layout(
            template="plotly_dark", paper_bgcolor="#0B0B0F", plot_bgcolor="#0f0f16",
            height=300, margin=dict(l=0,r=0,t=10,b=0),
            legend=dict(bgcolor="rgba(0,0,0,0)", font=dict(color="#A1A1AA")),
            xaxis=dict(title="Age", color="#3f3f46", gridcolor="rgba(255,255,255,0.04)"),
            yaxis=dict(color="#3f3f46", gridcolor="rgba(255,255,255,0.04)", tickprefix="€")
        )
        st.plotly_chart(fig_coast, use_container_width=True)

        if coast_reached_age:
            st.markdown(f"""
            <div style="background:#0f0f16;border:0.5px solid rgba(249,240,157,0.2);border-radius:10px;padding:18px 22px">
              <div style="font-size:13px;color:#F2F2F2;margin-bottom:4px">
                You reach Coast FI at approximately <strong style="color:#f9f09d">age {coast_reached_age}</strong>.
              </div>
              <div style="font-size:12px;color:#52525B">
                At that point you could stop contributing entirely and compounding alone would grow your portfolio to €{int(fn):,} by age {ret_age}.
              </div>
            </div>
            """, unsafe_allow_html=True)
        else:
            st.markdown("""
            <div style="background:#0f0f16;border:0.5px solid rgba(255,255,255,0.07);border-radius:10px;padding:18px 22px">
              <div style="font-size:13px;color:#71717A">Increase your monthly investment to reach Coast FI within your accumulation period.</div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown(divider("24px 0"), unsafe_allow_html=True)

        # ── RETIREMENT STRESS TEST ──
        st.markdown('<div style="font-family:\'Space Grotesk\',sans-serif;font-size:18px;font-weight:700;margin-bottom:4px">🧪 Retirement stress test</div>', unsafe_allow_html=True)
        st.markdown("""
        <div style="font-size:13px;color:#71717A;margin-bottom:6px">
          One of the biggest risks for retirees is a market downturn in the first years of withdrawal.
          Unlike during accumulation, early losses cannot be recovered by future contributions.
          This test shows how a market drop at different points in your retirement affects your portfolio.
        </div>
        <div style="font-size:12px;color:#3f3f46;margin-bottom:16px">
          Adjust the sliders to simulate different scenarios and see whether your plan survives.
        </div>
        """, unsafe_allow_html=True)

        def simulate_with_crash(monthly_inv, monthly_inc, years_acc, years_ret, ann_return, crash_pct, crash_year):
            capital, history, mr = 0, [], ann_return / 12
            for _ in range(years_acc * 12):
                capital = capital * (1 + mr) + monthly_inv
                history.append(capital)
            for m in range(years_ret * 12):
                y = m // 12
                if y == crash_year:
                    capital *= (1 - crash_pct)
                capital = capital * (1 + mr) - monthly_inc
                history.append(max(capital, 0))
                if capital <= 0: break
            return history

        col_sor1, col_sor2 = st.columns(2)
        with col_sor1:
            st.markdown('<div style="font-size:11px;color:#3f3f46;letter-spacing:1px;text-transform:uppercase;margin-bottom:6px">Market drop at retirement (%)</div>', unsafe_allow_html=True)
            crash_pct = st.slider("crash", 10, 50, 30, 5, format="-%d%%", label_visibility="collapsed") / 100
        with col_sor2:
            st.markdown('<div style="font-size:11px;color:#3f3f46;letter-spacing:1px;text-transform:uppercase;margin-bottom:6px">Drop occurs in retirement year</div>', unsafe_allow_html=True)
            crash_year = st.slider("crash_year", 1, min(10, yr), 1, label_visibility="collapsed")

        h_normal = simulate(mi, mc_inc, ya, yr, ar)[0]
        h_crash  = simulate_with_crash(mi, mc_inc, ya, yr, ar, crash_pct, crash_year)

        fig_sor = go.Figure()
        fig_sor.add_trace(go.Scatter(x=[i/12 for i in range(len(h_normal))], y=h_normal,
                                     mode="lines", name="No crash", line=dict(color="#7F77DD", width=2)))
        fig_sor.add_trace(go.Scatter(x=[i/12 for i in range(len(h_crash))], y=h_crash,
                                     mode="lines", name=f"{int(crash_pct*100)}% drop in year {crash_year}",
                                     line=dict(color="#E24B4A", width=2)))
        fig_sor.add_vline(x=ya, line_color="rgba(255,255,255,0.12)", line_dash="dash",
                          annotation_text="Retirement starts", annotation_font_color="#3f3f46")
        fig_sor.update_layout(
            template="plotly_dark", paper_bgcolor="#0B0B0F", plot_bgcolor="#0f0f16",
            height=300, margin=dict(l=0,r=0,t=10,b=0),
            legend=dict(bgcolor="rgba(0,0,0,0)", font=dict(color="#A1A1AA")),
            xaxis=dict(title="Years", color="#3f3f46", gridcolor="rgba(255,255,255,0.04)"),
            yaxis=dict(color="#3f3f46", gridcolor="rgba(255,255,255,0.04)", tickprefix="€")
        )
        st.plotly_chart(fig_sor, use_container_width=True)

        crash_survives = h_crash[-1] > 0 if h_crash else False
        sor_color = "#7F77DD" if crash_survives else "#E24B4A"
        st.markdown(f"""
        <div style="background:#0f0f16;border:0.5px solid {'rgba(127,119,221,0.25)' if crash_survives else 'rgba(226,75,74,0.25)'};border-radius:10px;padding:18px 22px">
          <div style="font-size:13px;font-weight:500;color:{sor_color};margin-bottom:6px">
            {'✓ Your plan survives this scenario.' if crash_survives else '✗ Your plan does not survive this scenario.'}
          </div>
          <div style="font-size:12px;color:#52525B;line-height:1.65">
            {'A {int(crash_pct*100)}% market drop in retirement year {crash_year} reduces your portfolio significantly, but your plan remains sustainable. Maintaining 2-3 years of expenses in cash or short-term bonds at retirement helps absorb this kind of shock without selling equities at the worst moment.'.format(int(crash_pct*100), crash_year) if crash_survives else 'This scenario depletes your portfolio before the end of retirement. Consider: (1) increasing your monthly investment now, (2) targeting a lower withdrawal rate, or (3) building a cash buffer of 2-3 years of expenses before retiring.'}
          </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown(divider("24px 0"), unsafe_allow_html=True)

        # ── HEATMAP ──
        st.markdown('<div style="font-family:\'Space Grotesk\',sans-serif;font-size:18px;font-weight:700;margin-bottom:4px">🌡️ Success probability heatmap</div>', unsafe_allow_html=True)
        st.markdown("<div style='font-size:13px;color:#71717A;margin-bottom:14px'>Probability of success across different withdrawal rates and retirement durations.</div>", unsafe_allow_html=True)

        swr_range = [0.03, 0.035, 0.04, 0.045, 0.05]
        dur_range = [15, 20, 25, 30, 35]
        heatmap_z = []
        for swr_val in swr_range:
            row = []
            for dur_val in dur_range:
                monthly_draw = (max_cap * swr_val) / 12
                ok = sum(
                    simulate(mi, monthly_draw, ya, dur_val, ar + np.random.normal(0, 0.015))[1]
                    for _ in range(200)
                )
                row.append(round(ok / 200 * 100))
            heatmap_z.append(row)

        fig_heat = go.Figure(data=go.Heatmap(
            z=heatmap_z,
            x=[f"{d} yrs" for d in dur_range],
            y=[f"{int(s*100)}%" for s in swr_range],
            colorscale=[[0, "#E24B4A"], [0.5, "#EF9F27"], [1, "#7F77DD"]],
            text=[[f"{v}%" for v in row] for row in heatmap_z],
            texttemplate="%{text}",
            showscale=True,
            zmin=0, zmax=100
        ))
        fig_heat.update_layout(
            template="plotly_dark", paper_bgcolor="#0B0B0F", plot_bgcolor="#0f0f16",
            height=280, margin=dict(l=0,r=0,t=10,b=0),
            xaxis=dict(title="Retirement duration", color="#A1A1AA"),
            yaxis=dict(title="Withdrawal rate", color="#A1A1AA")
        )
        st.plotly_chart(fig_heat, use_container_width=True)
        st.markdown("<div style='font-size:11px;color:#3f3f46;margin-top:-8px'>Based on 200 Monte Carlo simulations per cell using your accumulated portfolio at retirement.</div>", unsafe_allow_html=True)

        st.markdown(divider("24px 0"), unsafe_allow_html=True)

        # ── EXCEL EXPORT ──
        st.markdown('<div style="font-family:\'Space Grotesk\',sans-serif;font-size:18px;font-weight:700;margin-bottom:4px">📊 Excel projection</div>', unsafe_allow_html=True)
        st.markdown("<div style='font-size:13px;color:#71717A;margin-bottom:14px'>Your complete year-by-year projection as a downloadable Excel file.</div>", unsafe_allow_html=True)

        def generate_excel(inp_data, history_data, fn_val, score_val):
            import io as _io
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter
            except ImportError:
                return None

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Projection"

            dark   = "FF0B0B0F"
            purple = "FF5A54C4"
            light  = "FFF2F2F2"
            muted  = "FF71717A"
            surface= "FF0f0f16"

            def cell_style(cell, bold=False, color="FFF2F2F2", bg=None, size=11, align="left"):
                cell.font = Font(name="Calibri", bold=bold, color=color, size=size)
                cell.alignment = Alignment(horizontal=align, vertical="center")
                if bg:
                    cell.fill = PatternFill("solid", fgColor=bg)

            ws.column_dimensions["A"].width = 8
            ws.column_dimensions["B"].width = 22
            ws.column_dimensions["C"].width = 22
            ws.column_dimensions["D"].width = 22
            ws.column_dimensions["E"].width = 22

            ws.append(["THE FREEDOM PROJECT (fi) — Financial Projection"])
            cell_style(ws["A1"], bold=True, color=purple, size=13)
            ws.merge_cells("A1:E1")
            ws.row_dimensions[1].height = 30

            ws.append([f"Generated for {inp_data.get('name','')} · {datetime.now().strftime('%B %d, %Y')}"])
            cell_style(ws["A2"], color=muted, size=10)
            ws.merge_cells("A2:E2")

            ws.append([])

            headers = ["Year", "Age", "Total Invested (€)", "Portfolio Value (€)", "Market Returns (€)"]
            ws.append(headers)
            for i, h in enumerate(headers, 1):
                c = ws.cell(row=4, column=i)
                cell_style(c, bold=True, color=light, bg="FF26215C", size=11, align="center")
            ws.row_dimensions[4].height = 22

            acc_years = inp_data.get("years_accumulation", 25)
            monthly   = inp_data.get("monthly_investment", 300)
            ret_age_  = inp_data.get("current_age", 30) + acc_years

            for y in range(1, acc_years + 1):
                idx    = min(y * 12 - 1, len(history_data) - 1)
                pval   = history_data[idx] if idx < len(history_data) else 0
                tcontr = monthly * 12 * y
                mret   = max(0, pval - tcontr)
                cur_age = inp_data.get("current_age", 30) + y
                bg_row = "FF0f0f16" if y % 2 == 0 else "FF141420"
                row = [y, cur_age, int(tcontr), int(pval), int(mret)]
                ws.append(row)
                for i, val in enumerate(row, 1):
                    c = ws.cell(row=4 + y, column=i)
                    col_ = "FFf9f09d" if i == 4 else ("FF7F77DD" if i == 5 else light)
                    cell_style(c, color=col_, bg=bg_row, align="center")
                ws.row_dimensions[4 + y].height = 18

            ws.append([])
            summary_row = 4 + acc_years + 2
            ws.append(["Summary", "", "", "", ""])
            cell_style(ws.cell(row=summary_row, column=1), bold=True, color=purple, size=12)

            summaries = [
                ("Freedom Number", f"€{int(fn_val):,}"),
                ("Readiness Score", f"{int(score_val)}/100"),
                ("Total Invested", f"€{int(monthly * 12 * acc_years):,}"),
                ("Final Portfolio", f"€{int(history_data[acc_years*12-1] if len(history_data) >= acc_years*12 else 0):,}"),
            ]
            for i, (k, v) in enumerate(summaries):
                r = summary_row + 1 + i
                ws.cell(row=r, column=1, value=k)
                ws.cell(row=r, column=2, value=v)
                cell_style(ws.cell(row=r, column=1), color=muted, size=10)
                cell_style(ws.cell(row=r, column=2), bold=True, color=light, size=11)

            buf = _io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf

        excel_buf = generate_excel(inp, history, fn, score)
        if excel_buf:
            col_xl = st.columns([2, 1, 2])
            with col_xl[1]:
                st.download_button(
                    label="⬇ Download Excel",
                    data=excel_buf,
                    file_name=f"freedom_project_projection_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            st.info("Install openpyxl to enable Excel export: pip install openpyxl")
        st.markdown('<div style="font-family:\'Space Grotesk\',sans-serif;font-size:18px;font-weight:700;margin-bottom:4px">📄 Full Pro report</div>', unsafe_allow_html=True)
        st.markdown("<div style='font-size:13px;color:#71717A;margin-bottom:20px'>Comprehensive branded PDF including all Pro analysis: Monte Carlo, stress test, Coast FI, heatmap and year-by-year projection.</div>", unsafe_allow_html=True)

        pro_inp = {**inp,
                   "name":    st.session_state.inputs.get("pro_name", st.session_state.pro_email),
                   "email":   st.session_state.pro_email,
                   "country": inp.get("country", "—")}
        pro_pdf = generate_pdf(pro_inp, score, inv_needed, max_cap, target_cap, fn, fi_age, survives, gap,
                               mc_prob=mc_prob, coast_age=coast_reached_age,
                               contrib_pct=final_contrib_pct, returns_pct=final_returns_pct,
                               total_contrib=int(contrib_vals[-1]), total_returns=int(returns_vals[-1]))

        col_pro_dl = st.columns([2, 1, 2])
        with col_pro_dl[1]:
            st.download_button(
                label="⬇ Download Pro report",
                data=pro_pdf,
                file_name=f"freedom_project_pro_{datetime.now().strftime('%Y%m%d')}.pdf",
                mime="application/pdf"
            )

    # ── RESTART ──
    st.markdown(divider("32px 0"), unsafe_allow_html=True)
    col_r = st.columns([2, 1, 2])
    with col_r[1]:
        if st.button("← Start over"):
            for k in ["step", "inputs", "results", "show_pdf_gate", "pdf_unlocked", "pro_unlocked", "pro_email", "pro_gate_open"]:
                st.session_state[k] = 0 if k == "step" else ({} if k == "inputs" else ("" if k == "pro_email" else (False if k != "results" else None)))
            st.rerun()
